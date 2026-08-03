"""Tests for the Excel import / reconciliation pipeline (Unit 2)."""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import pytest

FIXTURES = Path(__file__).parent / "fixtures"
SAMPLE_XLSX = FIXTURES / "collection_export_sample.xlsx"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def make_cache(tmp_path: Path):
    from locg.collection_cache import CollectionCache
    return CollectionCache(
        path=tmp_path / "collection.json",
        lock_path=tmp_path / "collection.lock",
        audit_path=tmp_path / "import-history.jsonl",
    )


def make_agent_win_row(
    publisher: str = "Marvel",
    series: str = "Amazing Spider-Man",
    full_title: str = "Amazing Spider-Man #300",
    release_date: str = "1988-05-10",
    needs_manual_series: bool = False,
    needs_manual_variant: bool = False,
    gixen_item_id: str | None = "42",
    pushed: str | None = None,
) -> dict[str, Any]:
    return {
        "publisher_name": publisher,
        "series_name": series,
        "full_title": full_title,
        "release_date": release_date,
        "in_collection": 1,
        "in_wish_list": 0,
        "marked_read": 0,
        "my_rating": None,
        "media_format": "Print",
        "price_paid": None,
        "date_purchased": None,
        "condition": None,
        "notes": None,
        "tags": None,
        "storage_box": None,
        "owner": None,
        "purchase_store": None,
        "signature": None,
        "slabbing": None,
        "grading": None,
        "grading_company": None,
        "local_added_at": "2024-01-01T00:00:00.000000Z",
        "local_added_seq": 1,
        "pushed_to_locg_at": pushed,
        "last_seen_in_export_at": None,
        "source": "agent_win",
        "needs_manual_variant": needs_manual_variant,
        "needs_manual_series_canonical": needs_manual_series,
        "metron_id": None,
        "gixen_item_id": gixen_item_id,
        "previous_full_title": None,
    }


# ---------------------------------------------------------------------------
# parse_xlsx
# ---------------------------------------------------------------------------

def test_parse_xlsx_row_count():
    """parse_xlsx returns rows; count == max_row - 1 (minus header)."""
    from locg.collection_io import parse_xlsx
    rows = parse_xlsx(SAMPLE_XLSX)
    assert len(rows) > 0
    # Sample file has 2353 data rows (2354 including header)
    assert len(rows) == 2353


def test_parse_xlsx_row_shape():
    """Every parsed row has all 21 LOCG column keys."""
    from locg.collection_io import parse_xlsx
    from locg.collection_cache import LOCG_COLUMNS
    rows = parse_xlsx(SAMPLE_XLSX)
    for row in rows[:5]:
        for col in LOCG_COLUMNS:
            assert col in row, f"Missing column: {col}"


def test_parse_xlsx_first_row_values():
    """First data row matches known fixture values."""
    from locg.collection_io import parse_xlsx
    rows = parse_xlsx(SAMPLE_XLSX)
    first = rows[0]
    assert first["publisher_name"] == "Image Comics"
    assert first["series_name"] == "1963 (1993)"
    assert first["full_title"] == "1963 #6"
    assert first["in_collection"] == 1


def test_parse_xlsx_header_mismatch_raises(tmp_path):
    """A file with mismatched column headers raises RuntimeError before any row is read."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Publisher", "Wrong Column"])  # bad headers
    bad_path = tmp_path / "bad.xlsx"
    wb.save(bad_path)

    from locg.collection_io import parse_xlsx
    with pytest.raises(RuntimeError, match="header"):
        parse_xlsx(bad_path)


def test_parse_xlsx_file_too_large_rejected(tmp_path):
    """Files larger than 10 MB are rejected before parsing."""
    big_file = tmp_path / "big.xlsx"
    big_file.write_bytes(b"x" * (10 * 1024 * 1024 + 1))
    from locg.collection_io import parse_xlsx
    with pytest.raises(RuntimeError, match="10 MB"):
        parse_xlsx(big_file)


# ---------------------------------------------------------------------------
# import_xlsx — Phase 2 standard merge (happy paths)
# ---------------------------------------------------------------------------

def test_import_xlsx_populates_empty_cache(tmp_path):
    """Importing into an empty cache inserts all rows with source='locg_export'."""
    from locg.collection_io import import_xlsx
    cache = make_cache(tmp_path)
    result = import_xlsx(SAMPLE_XLSX, cache)
    assert result["added"] > 0
    assert result["updated"] == 0
    payload = cache.load()
    assert len(payload["comics"]) == result["added"]
    assert all(r["source"] == "locg_export" for r in payload["comics"])


def test_import_xlsx_sets_pushed_to_locg_at(tmp_path):
    """Rows imported from LOCG have pushed_to_locg_at set (they're already in LOCG)."""
    from locg.collection_io import import_xlsx
    cache = make_cache(tmp_path)
    import_xlsx(SAMPLE_XLSX, cache)
    payload = cache.load()
    for row in payload["comics"]:
        assert row["pushed_to_locg_at"] is not None


def test_reimport_same_xlsx_unchanged(tmp_path):
    """Re-importing the same xlsx updates last_seen_in_export_at but adds no new rows."""
    from locg.collection_io import import_xlsx
    cache = make_cache(tmp_path)
    r1 = import_xlsx(SAMPLE_XLSX, cache)
    r2 = import_xlsx(SAMPLE_XLSX, cache)
    assert r2["added"] == 0
    assert r2["updated"] >= 0  # last_seen_in_export_at updated


def test_import_updates_last_full_import(tmp_path):
    """After import, last_full_import is set in the payload."""
    from locg.collection_io import import_xlsx
    cache = make_cache(tmp_path)
    import_xlsx(SAMPLE_XLSX, cache)
    payload = cache.load()
    assert payload["last_full_import"] is not None
    assert payload["last_import_source"] == str(SAMPLE_XLSX)


def test_import_builds_series_name_index(tmp_path):
    """After import, series_name_index is non-empty."""
    from locg.collection_io import import_xlsx
    cache = make_cache(tmp_path)
    import_xlsx(SAMPLE_XLSX, cache)
    payload = cache.load()
    assert len(payload["series_name_index"]) > 0


# ---------------------------------------------------------------------------
# import_xlsx — Phase 2: cache-only rows preserved
# ---------------------------------------------------------------------------

def test_cache_only_agent_win_survives_import(tmp_path):
    """agent_win rows not in the xlsx are preserved, not deleted (v1 preserves)."""
    from locg.collection_cache import CollectionCache
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    win_row = make_agent_win_row(
        publisher="DC",
        series="Batman (1940 - 2011)",
        full_title="Batman #999 (Not in Export)",
        release_date="2999-01-01",
    )

    def add_win(payload):
        payload["comics"].append(win_row)

    cache.apply(add_win, command="pre-import")
    import_xlsx(SAMPLE_XLSX, cache)

    payload = cache.load()
    titles = {r["full_title"] for r in payload["comics"]}
    assert "Batman #999 (Not in Export)" in titles


def test_pushed_not_in_export_possibly_removed_logged(tmp_path):
    """A pushed row not appearing in the re-export logs a 'possibly_removed' audit record."""
    from locg.collection_cache import CollectionCache
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    # Add a "pushed" agent_win row that won't appear in the xlsx
    ghost_row = make_agent_win_row(
        publisher="Nonexistent",
        series="Ghost Series (2099)",
        full_title="Ghost #1",
        release_date="2099-01-01",
        pushed="2024-01-01T00:00:00Z",
    )

    def add_ghost(payload):
        payload["comics"].append(ghost_row)

    cache.apply(add_ghost, command="pre-import")
    result = import_xlsx(SAMPLE_XLSX, cache)

    # The row must still be in the cache
    payload = cache.load()
    titles = {r["full_title"] for r in payload["comics"]}
    assert "Ghost #1" in titles

    # Audit log must have a possibly_removed record
    audit_lines = (tmp_path / "import-history.jsonl").read_text().strip().splitlines()
    audit_types = [json.loads(l)["type"] for l in audit_lines]
    assert "possibly_removed" in audit_types


# ---------------------------------------------------------------------------
# import_xlsx — Phase 1 reconciliation (R60)
# ---------------------------------------------------------------------------

def test_reconciliation_best_guess_row_resolved(tmp_path):
    """A best-guess agent_win row matched by reconciliation heuristic gets identity rewritten."""
    from locg.collection_cache import CollectionCache
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    # Add a row that matches what's in the xlsx — use a known series/issue from fixture
    # The fixture has "1963 #6" by Image Comics
    unresolved = make_agent_win_row(
        publisher="Image Comics",
        series="1963",  # bare name, no year — needs reconciliation
        full_title="1963 #6",
        release_date="1993-11-08",
        needs_manual_series=True,
    )
    unresolved["local_added_at"] = "2024-01-02T00:00:00.000000Z"
    unresolved["local_added_seq"] = 1

    def add_unresolved(payload):
        payload["comics"].append(unresolved)

    cache.apply(add_unresolved, command="pre-import")
    import_xlsx(SAMPLE_XLSX, cache)

    payload = cache.load()
    # Find the reconciled row
    matched = [r for r in payload["comics"] if r["full_title"] == "1963 #6"
               and r.get("gixen_item_id") == "42"]
    assert len(matched) == 1, "Reconciled row not found (expected exactly one)"
    assert matched[0]["needs_manual_series_canonical"] is False
    assert matched[0]["series_name"] == "1963 (1993)"  # LOCG canonical name
    # Tracking fields preserved
    assert matched[0]["gixen_item_id"] == "42"
    assert matched[0]["local_added_at"] == "2024-01-02T00:00:00.000000Z"


def test_reconciliation_vol_mismatch_not_reconciled(tmp_path):
    """A row with mismatching (Vol. N) annotation is NOT reconciled per R60.

    BUI-547: the row is deliberately NOT flagged needs_manual_series_canonical.
    A flagged row now gets its series RE-RESOLVED before scoring (that is the
    whole point of the flag — the stored name is a best guess), which would
    rewrite the "(Vol. 2)" under test and stop exercising R60 at all. An
    unflagged pending win still reaches the same scoring path via the BUI-122
    clause, so this keeps testing exactly what it says it tests.
    """
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    # Force a Vol mismatch: cache says "Amazing Spider-Man (Vol. 2)" but xlsx has Vol. 1
    bad_match = make_agent_win_row(
        publisher="Marvel Comics",
        series="Amazing Spider-Man (Vol. 2)",  # wrong vol
        full_title="Amazing Spider-Man #300",
        release_date="1988-05-10",
        needs_manual_series=False,
    )

    def add_row(payload):
        payload["comics"].append(bad_match)

    cache.apply(add_row, command="pre-import")
    result = import_xlsx(SAMPLE_XLSX, cache)

    payload = cache.load()
    unreconciled = [
        r for r in payload["comics"]
        if r.get("gixen_item_id") == "42"
    ]
    assert len(unreconciled) == 1
    assert unreconciled[0]["series_name"] == "Amazing Spider-Man (Vol. 2)", (
        "declared-volume conflict must not be rewritten"
    )
    assert unreconciled[0]["pushed_to_locg_at"] is None, "row stays pending"
    assert result["reconciled"] == 0


# ---------------------------------------------------------------------------
# import_xlsx — BUI-122: year-tolerant reconciliation for unflagged pending
# agent_win rows (LOCG canonicalizes Release Date on re-export, breaking the
# Phase-2 exact identity match). Uses a deterministic in-test XLSX rather than
# the golden fixture so the date-shift / ambiguity scenarios are exact.
# ---------------------------------------------------------------------------

def _build_export_xlsx(path: Path, rows: list[dict[str, Any]]):
    """Build a 21-column LOCG-format XLSX from row dicts.

    Each row dict needs publisher/series/full_title/release_date; in_collection
    defaults to 1, in_wish_list to 0.
    """
    import openpyxl
    from locg.collection_io import LOCG_XLSX_HEADERS

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(list(LOCG_XLSX_HEADERS))
    for r in rows:
        ws.append([
            r["publisher"], r["series"], r["full_title"], r["release_date"],
            r.get("in_collection", 1), r.get("in_wish_list", 0), 0, None,
            "Print", r.get("price_paid"), r.get("date_purchased"),
            None, None, None, None, None, None,
            None, None, None, None,
        ])
    wb.save(path)


# ---------------------------------------------------------------------------
# parse_xlsx — cell type coercion (BUI-469): openpyxl returns raw cell values
# with no type coercion, so a date-formatted cell arrives as a datetime and a
# text-formatted boolean/count cell arrives as a str. Both classes share one
# root cause and are fixed by the same coercion step in parse_xlsx.
# ---------------------------------------------------------------------------

def test_parse_xlsx_coerces_datetime_release_date_cell_to_iso_string(tmp_path):
    """A date-formatted 'Release Date' cell comes back from openpyxl as a
    datetime; parse_xlsx must normalize it to the same 'YYYY-MM-DD' string
    shape every other release_date in the codebase uses."""
    from datetime import datetime as dt
    from locg.collection_io import parse_xlsx

    xlsx = tmp_path / "dated.xlsx"
    _build_export_xlsx(xlsx, [{
        "publisher": "DC Comics", "series": "Watchmen (1986 - 1987)",
        "full_title": "Watchmen #1", "release_date": dt(1986, 9, 2),
    }])

    rows = parse_xlsx(xlsx)
    assert rows[0]["release_date"] == "1986-09-02"
    assert isinstance(rows[0]["release_date"], str)


def test_parse_xlsx_coerces_text_in_collection_to_int_preserving_count(tmp_path):
    """A text-formatted 'In Collection' cell must coerce to int, never bool —
    bool("0") is True, and in_collection is a copies-owned count (0/1/2+),
    not a flag, so the distinct values must survive coercion."""
    from locg.collection_io import parse_xlsx

    xlsx = tmp_path / "text_counts.xlsx"
    _build_export_xlsx(xlsx, [
        {"publisher": "Marvel", "series": "Daredevil", "full_title": "Daredevil #1",
         "release_date": "1964-04-01", "in_collection": "0"},
        {"publisher": "Marvel", "series": "Daredevil", "full_title": "Daredevil #2",
         "release_date": "1964-06-01", "in_collection": "2"},
    ])

    rows = parse_xlsx(xlsx)
    by_title = {r["full_title"]: r for r in rows}

    assert by_title["Daredevil #1"]["in_collection"] == 0
    assert isinstance(by_title["Daredevil #1"]["in_collection"], int)
    assert by_title["Daredevil #2"]["in_collection"] == 2, \
        "copy count must survive coercion, not collapse to a bool"
    assert isinstance(by_title["Daredevil #2"]["in_collection"], int)


def test_coerce_count_cell_handles_nan_and_inf_without_raising():
    """A stray NaN/Infinity cell (e.g. a formula error openpyxl surfaces as a
    float) must default to 0, not raise. int(float("nan")) raises ValueError
    and int(float("inf")) raises OverflowError — both must be caught."""
    from locg.collection_io import _coerce_count_cell

    assert _coerce_count_cell(float("nan")) == 0
    assert _coerce_count_cell(float("inf")) == 0
    assert _coerce_count_cell(float("-inf")) == 0


def test_parse_xlsx_warns_on_unparseable_in_collection_cell(tmp_path, caplog):
    """An unparseable 'In Collection' cell must default to 0 (never raise) but
    log a warning — silently reading a garbled ownership cell as "not owned"
    is the R11-dangerous direction (a hidden duplicate-buy risk), so the
    anomaly must stay visible rather than disappear."""
    import logging
    from locg.collection_io import parse_xlsx

    xlsx = tmp_path / "garbled.xlsx"
    _build_export_xlsx(xlsx, [{
        "publisher": "Marvel", "series": "Daredevil", "full_title": "Daredevil #1",
        "release_date": "1964-04-01", "in_collection": "N/A",
    }])

    with caplog.at_level(logging.WARNING, logger="locg"):
        rows = parse_xlsx(xlsx)

    assert rows[0]["in_collection"] == 0
    assert isinstance(rows[0]["in_collection"], int)
    assert any("in_collection" in r.message and "N/A" in r.message for r in caplog.records)


def test_import_xlsx_with_date_formatted_release_date_does_not_raise(tmp_path):
    """BUI-469 acceptance: reproduces the exact pre-fix crash site.
    _reconcile_score's year compare runs `(row.get("release_date") or "")[:4]`
    on the xlsx row, which raises TypeError on a raw datetime — and it runs
    first, during reconciliation of an unflagged pending agent_win row."""
    from datetime import datetime as dt
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    win = make_agent_win_row(
        publisher="Marvel Comics",
        series="Amazing Spider-Man",
        full_title="Amazing Spider-Man #309",
        release_date="1988-01-01",  # agent stamped Jan 1
        needs_manual_series=False,  # unflagged pending path (BUI-122)
        pushed=None,
    )

    def add_win(payload):
        payload["comics"].append(win)

    cache.apply(add_win, command="pre-import")

    xlsx = tmp_path / "reexport.xlsx"
    _build_export_xlsx(xlsx, [{
        "publisher": "Marvel Comics", "series": "Amazing Spider-Man",
        "full_title": "Amazing Spider-Man #309",
        "release_date": dt(1988, 2, 11),  # date-formatted cell, not text
    }])

    result = import_xlsx(xlsx, cache)  # must not raise
    payload = cache.load()

    row = next(r for r in payload["comics"] if r["full_title"] == "Amazing Spider-Man #309")
    assert row["release_date"] == "1988-02-11"
    assert result["reconciled"] >= 1


def test_import_xlsx_text_in_collection_cell_correct_on_new_insert(tmp_path):
    """BUI-469 acceptance: a text-formatted 'In Collection' cell must read
    with correct ownership EVERYWHERE, not just the _is_owned branch BUI-462
    hardened. A brand-new row (Phase 2's plain-insert path, which never calls
    _is_owned) must still land a real int — the string "0" would read truthy
    (bool("0") is True) at every later touch point that checks ownership."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    xlsx = tmp_path / "export.xlsx"
    _build_export_xlsx(xlsx, [{
        "publisher": "DC Comics", "series": "Batman", "full_title": "Batman #1",
        "release_date": "1940-04-25", "in_collection": "0",  # text, not owned
    }])

    import_xlsx(xlsx, cache)
    payload = cache.load()
    row = next(r for r in payload["comics"] if r["full_title"] == "Batman #1")

    assert row["in_collection"] == 0
    assert isinstance(row["in_collection"], int)
    assert bool(row["in_collection"]) is False


def test_import_xlsx_text_in_collection_cell_preserves_copy_count_on_update(tmp_path):
    """A text-formatted multi-copy count ('2') must survive the standard-merge
    update path (_apply_locg_columns_held) as the int 2, not collapse to a
    bool — losing the 0/1/2+ distinction would misreport how many copies are
    actually owned."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    existing = _make_export_row(
        publisher="Marvel Comics", series="Daredevil", full_title="Daredevil #1",
        release_date="1964-04-01", in_collection=0, in_wish_list=1,
    )

    def seed(payload):
        payload["comics"].append(existing)

    cache.apply(seed, command="pre-import")

    xlsx = tmp_path / "reexport.xlsx"
    _build_export_xlsx(xlsx, [{
        "publisher": "Marvel Comics", "series": "Daredevil", "full_title": "Daredevil #1",
        "release_date": "1964-04-01", "in_collection": "2",  # text-formatted
    }])

    import_xlsx(xlsx, cache)
    payload = cache.load()
    row = next(r for r in payload["comics"] if r["full_title"] == "Daredevil #1")

    assert row["in_collection"] == 2, "copy count must survive, not collapse to a bool"
    assert isinstance(row["in_collection"], int)


def test_pending_agent_win_reconciled_on_within_year_date_shift(tmp_path):
    """BUI-122: an unflagged pending agent_win row whose export counterpart has
    the same year but a different Release Date reconciles instead of duplicating."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    win = make_agent_win_row(
        publisher="Marvel Comics",
        series="Amazing Spider-Man",
        full_title="Amazing Spider-Man #309",
        release_date="1988-01-01",  # agent stamped Jan 1
        needs_manual_series=False,  # NOT flagged — the new path
        pushed=None,                # pending
    )

    def add_win(payload):
        payload["comics"].append(win)

    cache.apply(add_win, command="pre-import")

    xlsx = tmp_path / "reexport.xlsx"
    _build_export_xlsx(xlsx, [{
        "publisher": "Marvel Comics",
        "series": "Amazing Spider-Man",
        "full_title": "Amazing Spider-Man #309",
        "release_date": "1988-02-11",  # LOCG canonicalized the date, same year
    }])

    result = import_xlsx(xlsx, cache)
    payload = cache.load()

    rows = [r for r in payload["comics"] if r["full_title"] == "Amazing Spider-Man #309"]
    assert len(rows) == 1, "must reconcile in place, not insert a duplicate"
    row = rows[0]
    assert row["pushed_to_locg_at"] is not None, "pending must clear"
    assert row["source"] == "locg_export"
    assert row["release_date"] == "1988-02-11", "identity rewritten to LOCG canonical"
    assert row["gixen_item_id"] == "42", "tracking field preserved"
    assert result["reconciled"] >= 1
    assert result["added"] == 0, "no new row inserted"


def test_pending_agent_win_with_null_publisher_reconciles(tmp_path):
    """BUI-122 production-faithful: agent_win rows are written with
    publisher_name=None (record-win has no publisher), while LOCG's export
    carries a canonical publisher. The row must still reconcile — a strict
    publisher compare would score it 0 and strand it pending forever."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    win = make_agent_win_row(
        publisher=None,  # as written by record-win in production
        series="Daredevil",
        full_title="Daredevil #181",
        release_date="1982-01-01",  # agent-stamped Jan 1
        needs_manual_series=False,
        pushed=None,
    )

    def add_win(payload):
        payload["comics"].append(win)

    cache.apply(add_win, command="pre-import")

    xlsx = tmp_path / "reexport.xlsx"
    _build_export_xlsx(xlsx, [{
        "publisher": "Marvel Comics",      # LOCG populates the publisher
        "series": "Daredevil",
        "full_title": "Daredevil #181",
        "release_date": "1982-04-10",      # ...and canonicalizes the date (same year)
    }])

    result = import_xlsx(xlsx, cache)
    payload = cache.load()

    rows = [r for r in payload["comics"] if r["full_title"] == "Daredevil #181"]
    assert len(rows) == 1, "null-publisher win must reconcile, not duplicate"
    row = rows[0]
    assert row["pushed_to_locg_at"] is not None, "pending must clear"
    assert row["publisher_name"] == "Marvel Comics", "identity adopts LOCG canonical publisher"
    assert result["reconciled"] >= 1
    assert result["added"] == 0


def test_pending_agent_win_exact_match_uses_phase2_not_reconciliation(tmp_path):
    """KTD-2: a pending agent_win row whose EXACT identity is in the export is
    handled by the Phase-2 standard merge, not routed through year-tolerant
    reconciliation — so a same-year variant in the export can't make it ambiguous."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    win = make_agent_win_row(
        publisher="Marvel Comics",
        series="Amazing Spider-Man",
        full_title="Amazing Spider-Man #300",
        release_date="1988-05-10",
        needs_manual_series=False,
        pushed=None,
    )

    def add_win(payload):
        payload["comics"].append(win)

    cache.apply(add_win, command="pre-import")

    xlsx = tmp_path / "reexport.xlsx"
    _build_export_xlsx(xlsx, [
        # Exact identity match for the win:
        {"publisher": "Marvel Comics", "series": "Amazing Spider-Man",
         "full_title": "Amazing Spider-Man #300", "release_date": "1988-05-10"},
        # A same-year variant that WOULD make reconciliation ambiguous if the
        # win were routed through it:
        {"publisher": "Marvel Comics", "series": "Amazing Spider-Man",
         "full_title": "Amazing Spider-Man #300 Newsstand", "release_date": "1988-05-10"},
    ])

    result = import_xlsx(xlsx, cache)
    payload = cache.load()

    win_rows = [r for r in payload["comics"]
                if r["full_title"] == "Amazing Spider-Man #300" and r.get("gixen_item_id") == "42"]
    assert len(win_rows) == 1
    row = win_rows[0]
    assert row["pushed_to_locg_at"] is not None, "exact match must clear pending via Phase 2"
    assert row["source"] == "locg_export"
    assert row["needs_manual_series_canonical"] is False
    assert result["updated"] >= 1, "matched via standard merge, not reconciliation"
    assert not result["warnings"], "exact-match primacy must avoid an ambiguity warning"


def test_pending_agent_win_collision_with_owned_row_auto_healed(tmp_path):
    """BUI-211: a pending agent_win win for a book already owned under an
    established locg_export identity must AUTO-HEAL — the redundant pending win
    is dropped, the owned row survives. (Pre-BUI-211 this was left pending +
    warned; that folded cleanup_duplicates.py class 1 into the reconciler.)

    This is also the 2026-06-23 production scenario: the two rows differ by
    publisher (None vs canonical) AND date (placeholder 1982-01-01 vs real)."""
    from locg.collection_io import import_xlsx, make_identity

    cache = make_cache(tmp_path)
    # Already-owned canonical row (from a prior import).
    owned = make_agent_win_row(
        publisher="Marvel Comics",
        series="Daredevil",
        full_title="Daredevil #181",
        release_date="1982-04-10",
        gixen_item_id=None,
        pushed="2024-01-01T00:00:00.000000Z",
    )
    owned["source"] = "locg_export"
    # A pending win for the SAME book, recorded with no publisher + fabricated date.
    win = make_agent_win_row(
        publisher=None, series="Daredevil", full_title="Daredevil #181",
        release_date="1982-01-01", gixen_item_id="99", pushed=None,
    )

    def add_rows(payload):
        payload["comics"].extend([owned, win])

    cache.apply(add_rows, command="pre-import")

    xlsx = tmp_path / "reexport.xlsx"
    # LOCG carries ONE canonical Daredevil #181 (it collapses the win into the
    # owned copy on its side).
    _build_export_xlsx(xlsx, [{
        "publisher": "Marvel Comics", "series": "Daredevil",
        "full_title": "Daredevil #181", "release_date": "1982-04-10",
    }])

    result = import_xlsx(xlsx, cache)
    payload = cache.load()

    dd = [r for r in payload["comics"] if r["full_title"] == "Daredevil #181"]
    # The redundant pending win ROW is gone (BUI-462 carries its gixen_item_id
    # onto the survivor, so identify the dropped row by its pending state).
    assert not any(
        r["source"] == "agent_win" and r["pushed_to_locg_at"] is None for r in dd
    ), "pending win must be healed away"
    # The established owned row survives, still owned.
    owned_row = next(r for r in dd if r["source"] == "locg_export")
    assert owned_row["gixen_item_id"] == "99", "the win's bid link is carried over"
    assert owned_row["in_collection"], "owned row survives owned"
    assert owned_row["pushed_to_locg_at"] is not None
    assert len(dd) == 1, "exactly one Daredevil #181 remains (no duplicate-identity pair)"
    # No duplicate-identity pair anywhere.
    idents = [make_identity(r) for r in payload["comics"]]
    assert len(idents) == len(set(idents)), "no duplicate-identity rows"
    assert result["auto_healed_duplicates"] >= 1
    assert result["reconciled"] == 0
    assert result["possibly_removed"] == 0, "a dedup heal is not a removal"
    assert not any("left pending" in w for w in result["warnings"]), \
        "healed case must not emit a leave-pending warning"


def test_two_pending_rows_collision_left_pending(tmp_path):
    """BUI-211 negative: when the collision target is NOT an established owned
    row (here two pending agent_win rows for the same book, neither pushed), the
    EXISTING leave-pending + warning behavior is preserved — no heal."""
    from locg.collection_io import import_xlsx, make_identity

    cache = make_cache(tmp_path)
    # An UNFLAGGED pending win whose EXACT identity matches the export (handled by
    # Phase 2 — it becomes the established target the second win collides with).
    target = make_agent_win_row(
        publisher="Marvel Comics", series="Daredevil",
        full_title="Daredevil #181", release_date="1982-04-10",
        gixen_item_id="11", pushed=None,
    )
    # A second pending win for the same book with a fabricated date → routed
    # through Phase-1 reconciliation, collides with `target` once Phase 2 would
    # claim the canonical identity. But Phase 1 runs first; the collision target
    # at that point is still the pending `target` row (source=agent_win, not
    # established), so it must be left pending — not healed.
    win = make_agent_win_row(
        publisher=None, series="Daredevil", full_title="Daredevil #181",
        release_date="1982-01-01", gixen_item_id="99", pushed=None,
    )

    def add_rows(payload):
        payload["comics"].extend([target, win])

    cache.apply(add_rows, command="pre-import")

    xlsx = tmp_path / "reexport.xlsx"
    _build_export_xlsx(xlsx, [{
        "publisher": "Marvel Comics", "series": "Daredevil",
        "full_title": "Daredevil #181", "release_date": "1982-04-10",
    }])

    result = import_xlsx(xlsx, cache)
    payload = cache.load()

    # The colliding pending win (id 99) is left pending (not healed away).
    win_row = next(r for r in payload["comics"] if r.get("gixen_item_id") == "99")
    assert win_row["pushed_to_locg_at"] is None, "win must stay pending — target not established-owned"
    assert result["auto_healed_duplicates"] == 0, "no heal against a non-established target"
    assert any("left pending" in w for w in result["warnings"]), "collision must still warn"
    # No duplicate-identity pair created.
    idents = [make_identity(r) for r in payload["comics"]]
    assert len(idents) == len(set(idents)), "no duplicate-identity rows"


def test_pending_agent_win_cross_year_not_reconciled(tmp_path):
    """A different YEAR (volume reboot) is not tolerated: the win stays pending and
    the export row inserts as a genuinely new row."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    win = make_agent_win_row(
        publisher="DC Comics",
        series="Action Comics Annual",
        full_title="Action Comics Annual #1",
        release_date="1987-01-01",
        needs_manual_series=False,
        pushed=None,
    )

    def add_win(payload):
        payload["comics"].append(win)

    cache.apply(add_win, command="pre-import")

    xlsx = tmp_path / "reexport.xlsx"
    _build_export_xlsx(xlsx, [{
        "publisher": "DC Comics", "series": "Action Comics Annual",
        "full_title": "Action Comics Annual #1", "release_date": "2012-06-01",  # 2012 reboot
    }])

    result = import_xlsx(xlsx, cache)
    payload = cache.load()

    rows = sorted(
        (r for r in payload["comics"] if r["full_title"] == "Action Comics Annual #1"),
        key=lambda r: r["release_date"],
    )
    assert len(rows) == 2, "different-year rows are distinct, not reconciled"
    original = next(r for r in rows if r.get("gixen_item_id") == "42")
    assert original["pushed_to_locg_at"] is None, "1987 win stays pending"
    assert result["added"] == 1, "2012 export row inserts as new"
    assert result["reconciled"] == 0


def test_clean_reconcile_no_collision_not_healed(tmp_path):
    """BUI-211 no-regression: a normal pending win that reconciles cleanly (its
    matched export identity collides with NOTHING) still merges as before — it is
    reconciled, not falsely healed away."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    win = make_agent_win_row(
        publisher="Marvel Comics",
        series="Daredevil",
        full_title="Daredevil #181",
        release_date="1982-01-01",  # placeholder; export canonicalizes (same year)
        needs_manual_series=False,
        gixen_item_id="77",
        pushed=None,
    )

    def add_win(payload):
        payload["comics"].append(win)

    cache.apply(add_win, command="pre-import")

    xlsx = tmp_path / "reexport.xlsx"
    _build_export_xlsx(xlsx, [{
        "publisher": "Marvel Comics", "series": "Daredevil",
        "full_title": "Daredevil #181", "release_date": "1982-04-10",
    }])

    result = import_xlsx(xlsx, cache)
    payload = cache.load()

    rows = [r for r in payload["comics"] if r["full_title"] == "Daredevil #181"]
    assert len(rows) == 1, "reconciled in place, not duplicated"
    row = rows[0]
    assert row["gixen_item_id"] == "77", "the win row survived (was reconciled, not dropped)"
    assert row["pushed_to_locg_at"] is not None, "pending cleared"
    assert row["source"] == "locg_export"
    assert result["reconciled"] >= 1
    assert result["auto_healed_duplicates"] == 0, "no collision → no heal"


# ---------------------------------------------------------------------------
# BUI-462: wish-twin reconciliation (extends the BUI-211 auto-heal)
# ---------------------------------------------------------------------------

def _make_export_row(
    *,
    publisher: str,
    series: str,
    full_title: str,
    release_date: str,
    in_collection: int,
    in_wish_list: int,
) -> dict[str, Any]:
    """A settled locg_export cache row (from a prior import)."""
    row = make_agent_win_row(
        publisher=publisher,
        series=series,
        full_title=full_title,
        release_date=release_date,
        gixen_item_id=None,
        pushed="2024-01-01T00:00:00.000000Z",
    )
    row["source"] = "locg_export"
    row["in_collection"] = in_collection
    row["in_wish_list"] = in_wish_list
    return row


def _audit_records(tmp_path: Path) -> list[dict[str, Any]]:
    log = tmp_path / "import-history.jsonl"
    if not log.exists():
        return []
    return [json.loads(ln) for ln in log.read_text().splitlines() if ln.strip()]


def test_wish_twin_plus_pending_win_auto_healed(tmp_path):
    """BUI-462 / the 2026-07-19 sync: a wished-THEN-won book has THREE rows in
    play — the wish twin, the pending agent_win, and the incoming owned export
    row. Pre-BUI-462 the twin's *pre-import* in_collection=0 failed the
    established-owned test and all 27 such books were stranded as duplicates.

    The twin holds the export row's identity exactly, so Phase 2 will own it;
    the win must heal away, leaving a single owned row."""
    from locg.collection_io import import_xlsx, make_identity

    cache = make_cache(tmp_path)
    wish_twin = _make_export_row(
        publisher="Marvel Comics",
        series="Marvel Tales (1964 - 1994)",
        full_title="Marvel Tales #223",
        release_date="1989-02-14",
        in_collection=0,
        in_wish_list=1,
    )
    wish_twin["my_rating"] = 8  # a user-managed value that must survive
    # The win: no publisher (BUI-458), Jan-1 placeholder date (BUI-210).
    win = make_agent_win_row(
        publisher=None,
        series="Marvel Tales",
        full_title="Marvel Tales #223",
        release_date="1989-01-01",
        gixen_item_id="777",
        pushed=None,
    )

    def add_rows(payload):
        payload["comics"].extend([wish_twin, win])

    cache.apply(add_rows, command="pre-import")

    xlsx = tmp_path / "reexport.xlsx"
    # LOCG moved the book from the wish list into the collection on upload.
    _build_export_xlsx(xlsx, [{
        "publisher": "Marvel Comics",
        "series": "Marvel Tales (1964 - 1994)",
        "full_title": "Marvel Tales #223",
        "release_date": "1989-02-14",
        "in_collection": 1,
        "in_wish_list": 0,
    }])

    result = import_xlsx(xlsx, cache)
    payload = cache.load()

    rows = [r for r in payload["comics"] if r["full_title"] == "Marvel Tales #223"]
    assert len(rows) == 1, "exactly one row remains — no duplicate-identity pair"
    kept = rows[0]
    # The owned locg_export copy is the one kept.
    assert kept["source"] == "locg_export"
    assert kept["in_collection"], "kept row is the owned copy"
    assert kept["release_date"] == "1989-02-14", "LOCG's canonical identity kept"
    assert kept["my_rating"] is None, "the kept row is the twin, refreshed from the export"
    # The wish row settles from LOCG's own export, never from a local guess.
    assert kept["in_wish_list"] == 0, "LOCG un-wished it; the twin mirrors that"
    assert not any(
        r["source"] == "agent_win" and r["pushed_to_locg_at"] is None
        for r in payload["comics"]
    ), "the redundant pending win is healed away"
    assert kept["gixen_item_id"] == "777", "the win's bid link is carried, not lost"

    assert result["auto_healed_duplicates"] == 1
    assert result["possibly_removed"] == 0, "a dedup heal is not a removal"
    assert not any("left pending" in w for w in result["warnings"])
    idents = [make_identity(r) for r in payload["comics"]]
    assert len(idents) == len(set(idents)), "no duplicate-identity rows"

    # The drop is fully reconstructable from the append-only audit log.
    healed = [a for a in _audit_records(tmp_path)
              if a["type"] == "auto_healed_duplicate_win"]
    assert len(healed) == 1
    assert healed[0]["details"]["gixen_item_id"] == "777"
    assert healed[0]["details"]["dropped_identity"][3] == "1989-01-01"


def test_cross_volume_decoy_hold_survives_wish_twin_heal(tmp_path):
    """BUI-462 safety invariant: an intentional cross-volume decoy hold — a
    vintage grail wish-listed while a modern volume of the SAME masthead is
    owned — must be untouched by the heal.

    Both volumes normalize to the same series key, so the heal cannot be
    passing for free on a series-name difference."""
    from locg.collection_cache import _normalize_series_key
    from locg.collection_io import import_xlsx, make_identity

    assert (
        _normalize_series_key("Ghost Rider (1973 - 1983)")
        == _normalize_series_key("Ghost Rider (2022)")
    ), "precondition: the two volumes must collapse to one normalized key"

    cache = make_cache(tmp_path)
    decoy_hold = _make_export_row(
        publisher="Marvel Comics",
        series="Ghost Rider (1973 - 1983)",
        full_title="Ghost Rider #5",
        release_date="1974-04-02",
        in_collection=0,
        in_wish_list=1,
    )
    modern_owned = _make_export_row(
        publisher="Marvel Comics",
        series="Ghost Rider (2022)",
        full_title="Ghost Rider #5",
        release_date="2022-08-31",
        in_collection=1,
        in_wish_list=0,
    )
    # A pending win for the MODERN issue (placeholder date, no publisher).
    win = make_agent_win_row(
        publisher=None,
        series="Ghost Rider",
        full_title="Ghost Rider #5",
        release_date="2022-01-01",
        gixen_item_id="555",
        pushed=None,
    )

    def add_rows(payload):
        payload["comics"].extend([decoy_hold, modern_owned, win])

    cache.apply(add_rows, command="pre-import")

    xlsx = tmp_path / "reexport.xlsx"
    _build_export_xlsx(xlsx, [
        {"publisher": "Marvel Comics", "series": "Ghost Rider (1973 - 1983)",
         "full_title": "Ghost Rider #5", "release_date": "1974-04-02",
         "in_collection": 0, "in_wish_list": 1},
        {"publisher": "Marvel Comics", "series": "Ghost Rider (2022)",
         "full_title": "Ghost Rider #5", "release_date": "2022-08-31",
         "in_collection": 1, "in_wish_list": 0},
    ])

    result = import_xlsx(xlsx, cache)
    payload = cache.load()

    gr = [r for r in payload["comics"] if r["full_title"] == "Ghost Rider #5"]
    assert len(gr) == 2, "the two volumes stay distinct; only the win is dropped"
    hold = next(r for r in gr if r["series_name"] == "Ghost Rider (1973 - 1983)")
    assert hold["in_wish_list"] == 1, "the decoy hold is still wish-listed"
    assert hold["in_collection"] == 0, "the decoy hold is still not owned"
    assert hold["release_date"] == "1974-04-02"

    owned = next(r for r in gr if r["series_name"] == "Ghost Rider (2022)")
    assert owned["in_collection"], "the modern owned copy is the one kept"
    assert not any(
        r["source"] == "agent_win" and r["pushed_to_locg_at"] is None
        for r in payload["comics"]
    ), "the modern win healed against the modern volume"
    assert owned["gixen_item_id"] == "555", "healed against the MODERN row, not the hold"
    assert hold["gixen_item_id"] is None, "nothing was carried onto the decoy hold"
    assert result["auto_healed_duplicates"] == 1
    idents = [make_identity(r) for r in payload["comics"]]
    assert len(idents) == len(set(idents))


def test_dateless_win_cannot_clear_against_wrong_volume(tmp_path):
    """BUI-462 fail-closed era gate: `_reconcile_score`'s year compare fails
    OPEN when either side is dateless, so a dateless win matches ANY era of a
    masthead. That must never be enough to DROP the win as a duplicate of a
    different volume — here the only candidate is a vintage decoy hold that
    LOCG has just marked owned, and the win is really a modern issue."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    decoy_hold = _make_export_row(
        publisher="Marvel Comics",
        series="Ghost Rider (1973 - 1983)",
        full_title="Ghost Rider #5",
        release_date="1974-04-02",
        in_collection=0,
        in_wish_list=1,
    )
    win = make_agent_win_row(
        publisher=None,
        series="Ghost Rider",
        full_title="Ghost Rider #5",
        release_date="",  # dateless — the era evidence is gone
        gixen_item_id="556",
        pushed=None,
    )

    def add_rows(payload):
        payload["comics"].extend([decoy_hold, win])

    cache.apply(add_rows, command="pre-import")

    xlsx = tmp_path / "reexport.xlsx"
    # LOCG now reports the VINTAGE issue as owned — without the era gate this
    # would make the vintage row an owned collision target and delete the win.
    _build_export_xlsx(xlsx, [{
        "publisher": "Marvel Comics", "series": "Ghost Rider (1973 - 1983)",
        "full_title": "Ghost Rider #5", "release_date": "1974-04-02",
        "in_collection": 1, "in_wish_list": 0,
    }])

    result = import_xlsx(xlsx, cache)
    payload = cache.load()

    win_row = next(r for r in payload["comics"] if r.get("gixen_item_id") == "556")
    assert win_row["pushed_to_locg_at"] is None, "the win survives, still pending"
    assert result["auto_healed_duplicates"] == 0, "no heal without era evidence"
    assert any("left pending" in w for w in result["warnings"])
    # Prove the ERA gate is what declined it, not some earlier bail.
    reasons = [a["details"].get("reason") for a in _audit_records(tmp_path)
               if a["type"] == "ambiguous_reconciliation"]
    assert "heal_declined_win_has_no_year" in reasons


def test_wish_twin_not_healed_when_export_still_unowned(tmp_path):
    """BUI-462 negative: the twin only counts as owned because the INCOMING
    export row says so. A book still merely wish-listed on LOCG is not an owned
    collision target — keep the leave-pending behavior."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    wish_twin = _make_export_row(
        publisher="Marvel Comics",
        series="Marvel Tales (1964 - 1994)",
        full_title="Marvel Tales #223",
        release_date="1989-02-14",
        in_collection=0,
        in_wish_list=1,
    )
    win = make_agent_win_row(
        publisher=None,
        series="Marvel Tales",
        full_title="Marvel Tales #223",
        release_date="1989-01-01",
        gixen_item_id="778",
        pushed=None,
    )

    def add_rows(payload):
        payload["comics"].extend([wish_twin, win])

    cache.apply(add_rows, command="pre-import")

    xlsx = tmp_path / "reexport.xlsx"
    _build_export_xlsx(xlsx, [{
        "publisher": "Marvel Comics", "series": "Marvel Tales (1964 - 1994)",
        "full_title": "Marvel Tales #223", "release_date": "1989-02-14",
        "in_collection": 0, "in_wish_list": 1,
    }])

    result = import_xlsx(xlsx, cache)
    payload = cache.load()

    win_row = next(r for r in payload["comics"] if r.get("gixen_item_id") == "778")
    assert win_row["pushed_to_locg_at"] is None, "still pending — nothing owns it"
    assert result["auto_healed_duplicates"] == 0
    assert any("left pending" in w for w in result["warnings"])
    twin = next(r for r in payload["comics"] if r.get("gixen_item_id") is None
                and r["full_title"] == "Marvel Tales #223")
    assert twin["in_wish_list"] == 1, "the wish row is untouched"


def test_row_carrying_wish_state_is_never_dropped_by_the_heal(tmp_path):
    """BUI-462 structural invariant: the heal may only ever drop a row that
    carries NO wish state. Today record-win never writes `in_wish_list=1` on a
    win, so this is unreachable through the normal producers — it is pinned
    here so the invariant survives a future producer change rather than
    depending on one. This is what makes "the decoy holds cannot be deleted" a
    property of the drop itself, not of what happens to feed it."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    owned = _make_export_row(
        publisher="Marvel Comics",
        series="Marvel Tales (1964 - 1994)",
        full_title="Marvel Tales #223",
        release_date="1989-02-14",
        in_collection=1,
        in_wish_list=0,
    )
    win = make_agent_win_row(
        publisher=None,
        series="Marvel Tales",
        full_title="Marvel Tales #223",
        release_date="1989-01-01",
        gixen_item_id="779",
        pushed=None,
    )
    win["in_wish_list"] = 1  # carries wish state → structurally undroppable

    def add_rows(payload):
        payload["comics"].extend([owned, win])

    cache.apply(add_rows, command="pre-import")

    xlsx = tmp_path / "reexport.xlsx"
    _build_export_xlsx(xlsx, [{
        "publisher": "Marvel Comics", "series": "Marvel Tales (1964 - 1994)",
        "full_title": "Marvel Tales #223", "release_date": "1989-02-14",
        "in_collection": 1, "in_wish_list": 0,
    }])

    result = import_xlsx(xlsx, cache)
    payload = cache.load()

    assert any(r.get("gixen_item_id") == "779" for r in payload["comics"]), \
        "a wish-bearing row must never be dropped by the heal"
    assert result["auto_healed_duplicates"] == 0
    assert any("left pending" in w for w in result["warnings"])


def test_healed_win_purchase_provenance_survives_on_the_kept_row(tmp_path):
    """BUI-462: the heal must be a dedup, not a data loss. A wish twin has by
    definition never carried a purchase price, and Phase 2 blanks price_paid /
    date_purchased from the export — so the dropped win's local-only provenance
    has to land on the survivor (and the whole row in the audit log)."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    wish_twin = _make_export_row(
        publisher="Marvel Comics",
        series="Marvel Tales (1964 - 1994)",
        full_title="Marvel Tales #223",
        release_date="1989-02-14",
        in_collection=0,
        in_wish_list=1,
    )
    win = make_agent_win_row(
        publisher=None,
        series="Marvel Tales",
        full_title="Marvel Tales #223",
        release_date="1989-01-01",
        gixen_item_id="777",
        pushed=None,
    )
    win["price_paid"] = 850.0
    win["date_purchased"] = "2026-07-12"
    win["metron_id"] = 55512

    def add_rows(payload):
        payload["comics"].extend([wish_twin, win])

    cache.apply(add_rows, command="pre-import")

    xlsx = tmp_path / "reexport.xlsx"
    _build_export_xlsx(xlsx, [{
        "publisher": "Marvel Comics", "series": "Marvel Tales (1964 - 1994)",
        "full_title": "Marvel Tales #223", "release_date": "1989-02-14",
        "in_collection": 1, "in_wish_list": 0,
    }])

    result = import_xlsx(xlsx, cache)
    payload = cache.load()

    kept = next(r for r in payload["comics"] if r["full_title"] == "Marvel Tales #223")
    assert kept["price_paid"] == 850.0, "the $850 purchase price must survive the heal"
    assert kept["date_purchased"] == "2026-07-12"
    assert kept["gixen_item_id"] == "777", "the link back to the bid survives"
    assert kept["metron_id"] == 55512
    assert result["auto_healed_duplicates"] == 1

    # The full dropped row is in the append-only log, so a wrong drop is reversible.
    healed = next(a for a in _audit_records(tmp_path)
                  if a["type"] == "auto_healed_duplicate_win")
    assert healed["details"]["dropped_row"]["price_paid"] == 850.0
    # And the deletion is not a silent success.
    assert any("auto-healed away" in w for w in result["warnings"])


def test_carry_never_clobbers_a_value_the_kept_row_already_holds(tmp_path):
    """BUI-462: the provenance carry-over only fills empty fields — it must not
    overwrite a price LOCG (or a prior round-trip) already supplied."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    owned = _make_export_row(
        publisher="Marvel Comics",
        series="Marvel Tales (1964 - 1994)",
        full_title="Marvel Tales #223",
        release_date="1989-02-14",
        in_collection=1,
        in_wish_list=0,
    )
    win = make_agent_win_row(
        publisher=None, series="Marvel Tales", full_title="Marvel Tales #223",
        release_date="1989-01-01", gixen_item_id="777", pushed=None,
    )
    win["price_paid"] = 850.0

    def add_rows(payload):
        payload["comics"].extend([owned, win])

    cache.apply(add_rows, command="pre-import")

    xlsx = tmp_path / "reexport.xlsx"
    _build_export_xlsx(xlsx, [{
        "publisher": "Marvel Comics", "series": "Marvel Tales (1964 - 1994)",
        "full_title": "Marvel Tales #223", "release_date": "1989-02-14",
        "in_collection": 1, "in_wish_list": 0, "price_paid": 12.5,
    }])

    import_xlsx(xlsx, cache)
    payload = cache.load()

    kept = next(r for r in payload["comics"] if r["full_title"] == "Marvel Tales #223")
    assert kept["price_paid"] == 12.5, "LOCG's own value wins; the carry only fills gaps"


def test_healed_row_indices_retracted_so_no_export_row_is_lost(tmp_path):
    """BUI-462: a healed row's stale identity/partial index entries stay live
    through Phase 2 unless retracted — long enough for the R67 rename path to
    claim the doomed row as a rename target for an UNRELATED export row, which
    then vanishes with it. Every export row must survive the import."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    wish_twin = _make_export_row(
        publisher="Marvel Comics",
        series="Marvel Tales (1964 - 1994)",
        full_title="Marvel Tales #223",
        release_date="1989-02-14",
        in_collection=0,
        in_wish_list=1,
    )
    # Publisher + series + release_date here form the PARTIAL identity that the
    # rename path keys on; the #224 export row below shares it exactly.
    win = make_agent_win_row(
        publisher="Marvel Comics",
        series="Marvel Tales (1964 - 1994)",
        full_title="Marvel Tales #223",
        release_date="1989-01-01",
        gixen_item_id="777",
        pushed=None,
    )

    def add_rows(payload):
        payload["comics"].extend([wish_twin, win])

    cache.apply(add_rows, command="pre-import")

    xlsx = tmp_path / "reexport.xlsx"
    _build_export_xlsx(xlsx, [
        {"publisher": "Marvel Comics", "series": "Marvel Tales (1964 - 1994)",
         "full_title": "Marvel Tales #223", "release_date": "1989-02-14",
         "in_collection": 1, "in_wish_list": 0},
        # Shares (publisher, series, release_date) with the healed win.
        {"publisher": "Marvel Comics", "series": "Marvel Tales (1964 - 1994)",
         "full_title": "Marvel Tales #224", "release_date": "1989-01-01",
         "in_collection": 1, "in_wish_list": 0},
    ])

    import_xlsx(xlsx, cache)
    payload = cache.load()

    titles = {r["full_title"] for r in payload["comics"]}
    assert "Marvel Tales #223" in titles
    assert "Marvel Tales #224" in titles, \
        "an owned export row must never be consumed by a row that is then dropped"


def test_text_typed_in_collection_cell_cannot_authorize_a_drop(tmp_path):
    """BUI-462: `parse_xlsx` does no type coercion and `bool("0")` is True. An
    export row saying NOT owned must not authorize retiring a win."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    wish_twin = _make_export_row(
        publisher="Marvel Comics",
        series="Marvel Tales (1964 - 1994)",
        full_title="Marvel Tales #223",
        release_date="1989-02-14",
        in_collection=0,
        in_wish_list=1,
    )
    win = make_agent_win_row(
        publisher=None, series="Marvel Tales", full_title="Marvel Tales #223",
        release_date="1989-01-01", gixen_item_id="801", pushed=None,
    )

    def add_rows(payload):
        payload["comics"].extend([wish_twin, win])

    cache.apply(add_rows, command="pre-import")

    xlsx = tmp_path / "reexport.xlsx"
    _build_export_xlsx(xlsx, [{
        "publisher": "Marvel Comics", "series": "Marvel Tales (1964 - 1994)",
        "full_title": "Marvel Tales #223", "release_date": "1989-02-14",
        "in_collection": "0", "in_wish_list": "1",  # text-formatted cells
    }])

    result = import_xlsx(xlsx, cache)
    payload = cache.load()

    assert any(r.get("gixen_item_id") == "801" for r in payload["comics"]), \
        "a text '0' In Collection cell must not read as ownership"
    assert result["auto_healed_duplicates"] == 0


def test_dateless_trade_win_still_heals(tmp_path):
    """BUI-462 no-regression: `_reconcile_score`'s TPB/HC/OGN branch matches on
    the full_title itself and never compares years, so requiring a year there
    would newly strand dateless trade wins that BUI-211 healed."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    owned = _make_export_row(
        publisher="DC Comics",
        series="Watchmen (1986 - 1987)",
        full_title="Watchmen HC",
        release_date="2013-11-05",
        in_collection=1,
        in_wish_list=0,
    )
    win = make_agent_win_row(
        publisher=None, series="Watchmen", full_title="Watchmen HC",
        release_date="",  # trades routinely land dateless
        gixen_item_id="900", pushed=None,
    )

    def add_rows(payload):
        payload["comics"].extend([owned, win])

    cache.apply(add_rows, command="pre-import")

    xlsx = tmp_path / "reexport.xlsx"
    _build_export_xlsx(xlsx, [{
        "publisher": "DC Comics", "series": "Watchmen (1986 - 1987)",
        "full_title": "Watchmen HC", "release_date": "2013-11-05",
        "in_collection": 1, "in_wish_list": 0,
    }])

    result = import_xlsx(xlsx, cache)
    payload = cache.load()

    assert result["auto_healed_duplicates"] == 1, "no issue token → no volume ambiguity"
    assert not any(r.get("gixen_item_id") == "900" and r["source"] == "agent_win"
                   for r in payload["comics"])


def test_era_decline_warning_names_the_side_that_lacks_the_year(tmp_path):
    """BUI-462: telling the operator to backfill a date that is already present
    strands the row forever across every subsequent sync."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    wish_twin = _make_export_row(
        publisher="Marvel Comics",
        series="Marvel Tales (1964 - 1994)",
        full_title="Marvel Tales #223",
        release_date="",
        in_collection=0,
        in_wish_list=1,
    )
    win = make_agent_win_row(
        publisher=None, series="Marvel Tales", full_title="Marvel Tales #223",
        release_date="1989-02-14",  # the win's date is FINE
        gixen_item_id="902", pushed=None,
    )

    def add_rows(payload):
        payload["comics"].extend([wish_twin, win])

    cache.apply(add_rows, command="pre-import")

    xlsx = tmp_path / "reexport.xlsx"
    _build_export_xlsx(xlsx, [{
        "publisher": "Marvel Comics", "series": "Marvel Tales (1964 - 1994)",
        "full_title": "Marvel Tales #223", "release_date": "",
        "in_collection": 1, "in_wish_list": 0,
    }])

    result = import_xlsx(xlsx, cache)

    assert result["auto_healed_duplicates"] == 0
    warning = next(w for w in result["warnings"] if "left pending" in w)
    assert "incoming export row carries no release date" in warning
    assert "backfill its release_date" not in warning, \
        "must not send the operator after a date that is already correct"
    reasons = [a["details"].get("reason") for a in _audit_records(tmp_path)
               if a["type"] == "ambiguous_reconciliation"]
    assert "heal_declined_export_has_no_year" in reasons


def test_era_confirmed_requires_a_real_year_on_both_sides():
    """BUI-462: the heal's era gate fails CLOSED — the complement of
    `_reconcile_score`'s fail-open year compare."""
    from locg.collection_io import _era_confirmed

    def row(date):
        return {"release_date": date}

    assert _era_confirmed(row("1989-01-01"), row("1989-02-14"))
    assert not _era_confirmed(row(""), row("1974-04-02"))
    assert not _era_confirmed(row(None), row("1974-04-02"))
    assert not _era_confirmed(row("1989-02-14"), row(""))
    assert not _era_confirmed(row("n/a"), row("n/a")), "unparseable is not a year"
    assert not _era_confirmed(row("1989-02-14"), row("2022-08-31"))


# ---------------------------------------------------------------------------
# BUI-470: same-book test unified with record-win's own dedup pair
# (_dedup_era_compatible / _dedup_variant_compatible, BUI-267), and
# in_collection treated as a copy COUNT rather than a flag on heal.
# ---------------------------------------------------------------------------

def test_same_book_confirmed_rejects_variant_mismatch():
    """BUI-470: `_same_book_confirmed` must decline a base win against an
    owned Newsstand/Direct/Facsimile copy of the same issue+year — exactly
    what stops record-win's own dedup (_dedup_variant_compatible, BUI-267)
    from treating them as the same book.

    This is a DIRECT unit test of the gate rather than an end-to-end
    `import_xlsx` reproduction: `make_identity` includes the raw `full_title`
    verbatim, and `_reconcile_score` (which selects the single xlsx_row a
    flagged win is compared against before this gate ever runs) requires
    either an identical trailing issue token with nothing after it, or an
    EXACT case-insensitive full_title match — both already imply
    suffix-agreement between `cache_row` and `xlsx_row`, so a genuine
    variant/printing MISMATCH between those two specific rows can never reach
    this gate via `import_xlsx` today (verified empirically: `_reconcile_score`
    returns 0 for every base-vs-suffixed or suffix-vs-different-suffix pair
    tried). This test locks in the gate's own correctness as a defense-in-depth
    invariant — belt-and-suspenders parity with record-win's dedup, and a
    safety net if `_reconcile_score`'s matching is ever loosened later."""
    from locg.collection_io import _same_book_confirmed

    base_win = {
        "full_title": "Uncanny X-Men #201",
        "release_date": "1986-01-01",
    }
    owned_newsstand = {
        "full_title": "Uncanny X-Men #201 Newsstand Edition",
        "release_date": "1986-01-01",
    }
    confirmed, reason, detail = _same_book_confirmed(base_win, owned_newsstand)
    assert not confirmed
    assert reason == "heal_declined_variant_mismatch"
    assert "distinct editions" in detail

    # Same suffix on both sides is still the same book.
    same_newsstand = {
        "full_title": "Uncanny X-Men #201 Newsstand Edition",
        "release_date": "1986-01-01",
    }
    confirmed2, _reason2, _detail2 = _same_book_confirmed(owned_newsstand, same_newsstand)
    assert confirmed2, "identical suffixes on both sides must still confirm"


def test_pending_agent_win_collision_credits_genuine_second_copy(tmp_path):
    """BUI-470: `in_collection` is a copy COUNT, not a flag. A pending win
    that collides with a book ALREADY owned independently of this win (an
    established row that was owned before this win ever existed) represents
    a genuine second physical copy, not a redundant duplicate record of the
    same transition. The heal must credit it onto the survivor's
    in_collection rather than silently dropping it and trusting whatever
    count the export happens to still carry (which, absent a manual LOCG
    update, is still 1)."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    # Already-owned canonical row — established from a genuinely separate
    # purchase/import BEFORE this second win ever existed.
    owned = make_agent_win_row(
        publisher="Marvel Comics", series="Daredevil",
        full_title="Daredevil #181", release_date="1982-04-10",
        gixen_item_id=None, pushed="2024-01-01T00:00:00.000000Z",
    )
    owned["source"] = "locg_export"
    # A SEPARATE win for a second physical copy of the SAME book — record-win's
    # own dedup missed it (as in the BUI-211 production scenario: no publisher
    # + a placeholder date at write time), so it exists as a distinct pending
    # row with its own gixen_item_id (a distinct eBay purchase).
    second_copy_win = make_agent_win_row(
        publisher=None, series="Daredevil", full_title="Daredevil #181",
        release_date="1982-01-01", gixen_item_id="200", pushed=None,
    )

    def add_rows(payload):
        payload["comics"].extend([owned, second_copy_win])

    cache.apply(add_rows, command="pre-import")

    xlsx = tmp_path / "reexport.xlsx"
    # LOCG's own export still reports only ONE copy — the user hasn't (yet)
    # told LOCG about the second physical copy.
    _build_export_xlsx(xlsx, [{
        "publisher": "Marvel Comics", "series": "Daredevil",
        "full_title": "Daredevil #181", "release_date": "1982-04-10",
        "in_collection": 1,
    }])

    result = import_xlsx(xlsx, cache)
    payload = cache.load()

    dd = [r for r in payload["comics"] if r["full_title"] == "Daredevil #181"]
    assert len(dd) == 1, "still a single row for the identity — no duplicate-identity pair"
    kept = dd[0]
    assert kept["in_collection"] == 2, "the second copy must be credited, not dropped"
    assert kept["gixen_item_id"] == "200", "the win's bid link is still carried over"
    assert result["auto_healed_duplicates"] == 1
    assert result["second_copies_credited"] == 1

    credited = [a for a in _audit_records(tmp_path) if a["type"] == "second_copy_credited"]
    assert len(credited) == 1
    assert credited[0]["details"]["in_collection_before"] == 1
    assert credited[0]["details"]["in_collection_after"] == 2
    assert credited[0]["details"]["credited"] == 1
    assert any("credited as genuine extra copies" in w for w in result["warnings"])


def test_wish_twin_heal_still_not_credited_as_second_copy(tmp_path):
    """BUI-470 no-regression: the ORIGINAL wish-twin case (BUI-462) must NOT be
    credited as a second copy — the target was NOT independently owned before
    this win (in_collection=0, wish-only), so the fold IS the ownership
    transition itself. Crediting it on top of the export's own in_collection=1
    would double-count a book that was only ever bought once."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    wish_twin = _make_export_row(
        publisher="Marvel Comics",
        series="Marvel Tales (1964 - 1994)",
        full_title="Marvel Tales #223",
        release_date="1989-02-14",
        in_collection=0,
        in_wish_list=1,
    )
    win = make_agent_win_row(
        publisher=None, series="Marvel Tales", full_title="Marvel Tales #223",
        release_date="1989-01-01", gixen_item_id="777", pushed=None,
    )

    def add_rows(payload):
        payload["comics"].extend([wish_twin, win])

    cache.apply(add_rows, command="pre-import")

    xlsx = tmp_path / "reexport.xlsx"
    _build_export_xlsx(xlsx, [{
        "publisher": "Marvel Comics", "series": "Marvel Tales (1964 - 1994)",
        "full_title": "Marvel Tales #223", "release_date": "1989-02-14",
        "in_collection": 1, "in_wish_list": 0,
    }])

    result = import_xlsx(xlsx, cache)
    payload = cache.load()

    kept = next(r for r in payload["comics"] if r["full_title"] == "Marvel Tales #223")
    assert kept["in_collection"] == 1, "wish-becomes-owned is ONE copy, not credited as a second"
    assert result["auto_healed_duplicates"] == 1
    assert result["second_copies_credited"] == 0
    assert not any("credited as genuine extra copies" in w for w in result["warnings"])


def test_pending_agent_win_ambiguous_left_pending(tmp_path):
    """When a pending agent_win row (no exact match) year-matches multiple export
    rows, it is left pending with an ambiguous_reconciliation audit/warning rather
    than guessing (duplicate-row policy: visible non-clear over silent wrong merge)."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    win = make_agent_win_row(
        publisher="Image Comics",
        series="Spawn",
        full_title="Spawn #300",
        release_date="2019-01-01",  # agent date; export will not carry this exact date
        needs_manual_series=False,
        pushed=None,
    )

    def add_win(payload):
        payload["comics"].append(win)

    cache.apply(add_win, command="pre-import")

    # Two printings: same publisher/series/issue-token/year, different dates.
    # Both score against the win (issue "300" + year 2019), so reconciliation is
    # ambiguous. (A variant whose Full Title doesn't end in "#300" wouldn't score,
    # since _reconcile_score requires an exact issue-token match.)
    xlsx = tmp_path / "reexport.xlsx"
    _build_export_xlsx(xlsx, [
        {"publisher": "Image Comics", "series": "Spawn",
         "full_title": "Spawn #300", "release_date": "2019-08-28"},
        {"publisher": "Image Comics", "series": "Spawn",
         "full_title": "Spawn #300", "release_date": "2019-11-13"},
    ])

    result = import_xlsx(xlsx, cache)
    payload = cache.load()

    win_rows = [r for r in payload["comics"] if r.get("gixen_item_id") == "42"]
    assert len(win_rows) == 1
    assert win_rows[0]["pushed_to_locg_at"] is None, "ambiguous match stays pending"
    assert result["reconciled"] == 0
    assert result["warnings"], "an ambiguity warning must be surfaced"
    audit_types = [
        json.loads(line)["type"]
        for line in (tmp_path / "import-history.jsonl").read_text().strip().splitlines()
    ]
    assert "ambiguous_reconciliation" in audit_types


# ---------------------------------------------------------------------------
# import_xlsx — renamed full_title persistence (R67)
# ---------------------------------------------------------------------------

def test_renamed_full_title_persists_previous(tmp_path):
    """When LOCG renames a full_title, previous_full_title is set for one cycle."""
    from locg.collection_cache import CollectionCache
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    # Seed cache with a row using the "old" title
    old_title = "1963 #6 (Old Name)"
    old_row = {
        "publisher_name": "Image Comics",
        "series_name": "1963 (1993)",
        "full_title": old_title,
        "release_date": "1993-11-08",
        "in_collection": 1,
        "in_wish_list": 0,
        "marked_read": 0,
        "my_rating": None,
        "media_format": "Print",
        "price_paid": None,
        "date_purchased": None,
        "condition": None,
        "notes": None,
        "tags": None,
        "storage_box": None,
        "owner": None,
        "purchase_store": None,
        "signature": None,
        "slabbing": None,
        "grading": None,
        "grading_company": None,
        "local_added_at": "2024-01-01T00:00:00.000000Z",
        "local_added_seq": 1,
        "pushed_to_locg_at": "2024-01-01T00:00:00Z",
        "last_seen_in_export_at": None,
        "source": "locg_export",
        "needs_manual_variant": False,
        "needs_manual_series_canonical": False,
        "metron_id": None,
        "gixen_item_id": None,
        "previous_full_title": None,
    }

    def seed(payload):
        payload["comics"].append(old_row)

    cache.apply(seed, command="seed")
    import_xlsx(SAMPLE_XLSX, cache)

    payload = cache.load()
    # The import should have matched by (publisher, series, release_date) and updated title
    renamed = [r for r in payload["comics"]
               if r.get("previous_full_title") == old_title]
    assert len(renamed) == 1, "Expected exactly one row with previous_full_title set"
    assert renamed[0]["full_title"] == "1963 #6"  # LOCG's canonical title


# ---------------------------------------------------------------------------
# import_xlsx — behavioral drift detection (F5)
# ---------------------------------------------------------------------------

def test_behavioral_drift_detected(tmp_path):
    """A changed user-managed column logs a behavioral_drift audit record."""
    from locg.collection_cache import CollectionCache
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    # Seed with a row that has a non-null my_rating that the import will clear
    row_with_rating = {
        "publisher_name": "Image Comics",
        "series_name": "1963 (1993)",
        "full_title": "1963 #6",
        "release_date": "1993-11-08",
        "in_collection": 1,
        "in_wish_list": 0,
        "marked_read": 0,
        "my_rating": 9.0,  # user set this; xlsx returns None → drift
        "media_format": "Print",
        "price_paid": None,
        "date_purchased": None,
        "condition": None,
        "notes": None,
        "tags": None,
        "storage_box": None,
        "owner": None,
        "purchase_store": None,
        "signature": None,
        "slabbing": None,
        "grading": None,
        "grading_company": None,
        "local_added_at": "2024-01-01T00:00:00.000000Z",
        "local_added_seq": 1,
        "pushed_to_locg_at": "2024-01-01T00:00:00Z",
        "last_seen_in_export_at": None,
        "source": "locg_export",
        "needs_manual_variant": False,
        "needs_manual_series_canonical": False,
        "metron_id": None,
        "gixen_item_id": None,
        "previous_full_title": None,
    }

    def seed(payload):
        payload["comics"].append(row_with_rating)

    cache.apply(seed, command="seed")
    import_xlsx(SAMPLE_XLSX, cache)

    audit_lines = (tmp_path / "import-history.jsonl").read_text().strip().splitlines()
    audit_types = [json.loads(l)["type"] for l in audit_lines]
    assert "behavioral_drift" in audit_types


# ---------------------------------------------------------------------------
# import_xlsx — series_name_index rebuilt from locg_export only
# ---------------------------------------------------------------------------

def test_series_name_index_after_import(tmp_path):
    """series_name_index is rebuilt from locg_export rows only after import."""
    from locg.collection_cache import CollectionCache
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    # Add an agent_win row — its series must NOT appear in the index
    win = make_agent_win_row(series="Totally Made Up Series (2099)")

    def add_win(payload):
        payload["comics"].append(win)

    cache.apply(add_win, command="pre-import")
    import_xlsx(SAMPLE_XLSX, cache)

    payload = cache.load()
    index = payload["series_name_index"]
    assert not any("totally made up" in k.lower() for k in index)
    # But fixture series should be present
    assert any("1963" in k for k in index)


# ---------------------------------------------------------------------------
# import_xlsx — crash recovery: migration_in_progress stays False
# ---------------------------------------------------------------------------

def test_import_never_leaves_migration_flag(tmp_path):
    """After a successful import, migration_in_progress must be False."""
    from locg.collection_io import import_xlsx
    cache = make_cache(tmp_path)
    import_xlsx(SAMPLE_XLSX, cache)
    payload = cache.load()
    assert payload["migration_in_progress"] is False


# ---------------------------------------------------------------------------
# import_xlsx — error paths
# ---------------------------------------------------------------------------

def test_import_nonexistent_file_raises(tmp_path):
    from locg.collection_io import import_xlsx
    cache = make_cache(tmp_path)
    with pytest.raises((RuntimeError, FileNotFoundError, OSError)):
        import_xlsx(tmp_path / "does_not_exist.xlsx", cache)


def test_import_bad_header_raises_before_merge(tmp_path):
    """Header mismatch must raise before any cache mutation."""
    import openpyxl
    from locg.collection_io import import_xlsx

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Wrong", "Headers", "Here"])
    bad_path = tmp_path / "bad_headers.xlsx"
    wb.save(bad_path)

    cache = make_cache(tmp_path)
    with pytest.raises(RuntimeError, match="header"):
        import_xlsx(bad_path, cache)

    # Cache should be untouched — no file created
    assert not (tmp_path / "collection.json").exists()


# ---------------------------------------------------------------------------
# Unit 3: generate_csv
# ---------------------------------------------------------------------------

RECIPE_CSV = FIXTURES / "locg_import_test_recipe.csv"


def _make_ready_row(
    publisher: str = "Marvel Comics",
    series: str = "The Amazing Spider-Man (Vol. 1) (1962 - 1998)",
    full_title: str = "The Amazing Spider-Man #84",
    release_date: str = "1970-05-01",
    price_paid: Any = 27.86,
    date_purchased: Any = "2026-05-22",
) -> dict[str, Any]:
    """Build a minimal agent_win row suitable for CSV export."""
    return {
        "publisher_name": publisher,
        "series_name": series,
        "full_title": full_title,
        "release_date": release_date,
        "in_collection": 1,
        "in_wish_list": 0,
        "marked_read": 0,
        "my_rating": None,
        "media_format": "Print",
        "price_paid": price_paid,
        "date_purchased": date_purchased,
        "condition": None,
        "notes": None,
        "tags": None,
        "storage_box": None,
        "owner": None,
        "purchase_store": "eBay",
        "signature": 0,
        "slabbing": 0,
        "grading": None,
        "grading_company": None,
        "local_added_at": "2026-05-22T10:00:00.000000Z",
        "local_added_seq": 1,
        "pushed_to_locg_at": None,
        "last_seen_in_export_at": None,
        "source": "agent_win",
        "needs_manual_variant": False,
        "needs_manual_series_canonical": False,
        "metron_id": None,
        "gixen_item_id": "12345",
        "previous_full_title": None,
    }


def test_generate_csv_wish_rows_default_raises(tmp_path):
    """BUI-208 machine gate: non-empty wish_rows with the default
    allow_uncollect=False refuses to write (would emit In Collection=0 rows
    that tell LOCG to DELETE owned titles)."""
    import pytest
    from locg.collection_io import generate_csv
    out = tmp_path / "out.csv"
    with pytest.raises(ValueError, match="In Collection=0"):
        generate_csv([_make_ready_row()], out, wish_rows=[_make_ready_row()])
    assert not out.exists()


def test_generate_csv_wish_rows_with_allow_uncollect_writes(tmp_path):
    """With allow_uncollect=True the wish rows are appended as In Collection=0,
    In Wish List=1 (the explicit owned-safe wish push)."""
    import csv
    from locg.collection_io import generate_csv
    out = tmp_path / "out.csv"
    generate_csv(
        [_make_ready_row(full_title="ASM #84")],
        out,
        wish_rows=[_make_ready_row(full_title="Saga #1")],
        allow_uncollect=True,
    )
    with open(out, newline="") as f:
        rows = list(csv.DictReader(f))
    by_title = {r["Full Title"]: r for r in rows}
    assert by_title["ASM #84"]["In Collection"] == "1"
    assert by_title["ASM #84"]["In Wish List"] == "0"
    assert by_title["Saga #1"]["In Collection"] == "0"
    assert by_title["Saga #1"]["In Wish List"] == "1"


def test_generate_csv_row_count(tmp_path):
    """10 ready rows produce a 10-row CSV (plus header)."""
    import csv
    from locg.collection_io import generate_csv
    rows = [_make_ready_row(full_title=f"ASM #{i}") for i in range(10)]
    out = tmp_path / "out.csv"
    generate_csv(rows, out)
    with open(out, newline="") as f:
        reader = list(csv.reader(f))
    assert len(reader) == 11  # 1 header + 10 data rows


def test_generate_csv_header_order(tmp_path):
    """CSV header matches the canonical 21-column LOCG order."""
    import csv
    from locg.collection_io import LOCG_XLSX_HEADERS, generate_csv
    generate_csv([_make_ready_row()], tmp_path / "out.csv")
    with open(tmp_path / "out.csv", newline="") as f:
        header = next(csv.reader(f))
    assert tuple(header) == LOCG_XLSX_HEADERS


def test_generate_csv_my_rating_blank_in_body(tmp_path):
    """My Rating column is present in header AND body as an empty string (R27)."""
    from locg.collection_io import generate_csv
    generate_csv([_make_ready_row()], tmp_path / "out.csv")
    raw = (tmp_path / "out.csv").read_text()
    lines = raw.splitlines()
    # Header must contain My Rating
    assert "My Rating" in lines[0]
    # Body row: the My Rating field position must be empty
    import csv, io
    reader = list(csv.reader(io.StringIO(raw)))
    header = reader[0]
    my_rating_idx = header.index("My Rating")
    body_row = reader[1]
    assert body_row[my_rating_idx] == ""


def test_generate_csv_omits_placeholder_release_date(tmp_path):
    """BUI-199 Cause 2: a placeholder-dated (YYYY-01-01) agent_win row exports
    with a BLANK Release Date, while a real-dated row keeps its date."""
    import csv, io
    from locg.collection_io import generate_csv

    placeholder = _make_ready_row(
        full_title="Placeholder Book #1", release_date="1988-01-01"
    )
    real = _make_ready_row(
        full_title="Real Book #1", release_date="1988-05-10"
    )
    out = tmp_path / "out.csv"
    generate_csv([placeholder, real], out)

    reader = list(csv.reader(io.StringIO(out.read_text())))
    header = reader[0]
    rd_idx = header.index("Release Date")
    ft_idx = header.index("Full Title")
    by_title = {row[ft_idx]: row for row in reader[1:]}

    assert by_title["Placeholder Book #1"][rd_idx] == ""
    assert by_title["Real Book #1"][rd_idx] == "1988-05-10"


def test_generate_csv_keeps_real_metron_jan1_date(tmp_path):
    """BUI-199 finding 5: a real Metron-sourced YYYY-01-01 cover_date (metron_id
    set) is KEPT on export; only a metron_id-less placeholder is blanked."""
    import csv, io
    from locg.collection_io import generate_csv

    metron_jan = _make_ready_row(
        full_title="Metron Jan Book #1", release_date="1988-01-01"
    )
    metron_jan["metron_id"] = 12345  # real Metron-backed date

    placeholder = _make_ready_row(
        full_title="Placeholder Book #1", release_date="1988-01-01"
    )
    placeholder["metron_id"] = None  # BUI-105 placeholder

    out = tmp_path / "out.csv"
    generate_csv([metron_jan, placeholder], out)

    reader = list(csv.reader(io.StringIO(out.read_text())))
    header = reader[0]
    rd_idx = header.index("Release Date")
    ft_idx = header.index("Full Title")
    by_title = {row[ft_idx]: row for row in reader[1:]}

    assert by_title["Metron Jan Book #1"][rd_idx] == "1988-01-01"
    assert by_title["Placeholder Book #1"][rd_idx] == ""


def test_generate_csv_keeps_placeholder_date_for_non_agent_win(tmp_path):
    """The placeholder-date omission is scoped to agent_win rows only — a
    locg_export row with a Jan-1 date keeps it (it is LOCG's real date)."""
    import csv, io
    from locg.collection_io import generate_csv

    row = _make_ready_row(full_title="Export Book #1", release_date="1988-01-01")
    row["source"] = "locg_export"
    out = tmp_path / "out.csv"
    generate_csv([row], out)

    reader = list(csv.reader(io.StringIO(out.read_text())))
    header = reader[0]
    rd_idx = header.index("Release Date")
    assert reader[1][rd_idx] == "1988-01-01"


def test_generate_csv_empty_queue_header_only(tmp_path):
    """Zero ready rows produces a CSV with only the header line."""
    import csv
    from locg.collection_io import generate_csv
    generate_csv([], tmp_path / "out.csv")
    with open(tmp_path / "out.csv", newline="") as f:
        rows = list(csv.reader(f))
    assert len(rows) == 1  # header only


def test_generate_csv_price_format(tmp_path):
    """Price Paid is formatted as NN.NN with no currency suffix."""
    import csv
    from locg.collection_io import generate_csv
    generate_csv([_make_ready_row(price_paid=27.8600001)], tmp_path / "out.csv")
    with open(tmp_path / "out.csv", newline="") as f:
        rows = list(csv.reader(f))
    header = rows[0]
    price_idx = header.index("Price Paid")
    assert rows[1][price_idx] == "27.86"


def test_generate_csv_negative_price_defaults_to_zero(tmp_path):
    """Negative price_paid defaults to 0.00 (R29)."""
    import csv
    from locg.collection_io import generate_csv
    generate_csv([_make_ready_row(price_paid=-5.0)], tmp_path / "out.csv")
    with open(tmp_path / "out.csv", newline="") as f:
        rows = list(csv.reader(f))
    price_idx = rows[0].index("Price Paid")
    assert rows[1][price_idx] == "0.00"


def test_generate_csv_missing_price_defaults_to_zero(tmp_path):
    """None price_paid defaults to 0.00."""
    import csv
    from locg.collection_io import generate_csv
    generate_csv([_make_ready_row(price_paid=None)], tmp_path / "out.csv")
    with open(tmp_path / "out.csv", newline="") as f:
        rows = list(csv.reader(f))
    price_idx = rows[0].index("Price Paid")
    assert rows[1][price_idx] == "0.00"


def test_generate_csv_date_iso_format(tmp_path):
    """Date Purchased is output as ISO date (YYYY-MM-DD) (R30)."""
    import csv
    from locg.collection_io import generate_csv
    generate_csv([_make_ready_row(date_purchased="2026-05-22")], tmp_path / "out.csv")
    with open(tmp_path / "out.csv", newline="") as f:
        rows = list(csv.reader(f))
    date_idx = rows[0].index("Date Purchased")
    assert rows[1][date_idx] == "2026-05-22"


def test_generate_csv_fixed_fields(tmp_path):
    """In Collection=1, In Wish List=0, Marked Read=0, Media Format=Print, Purchase Store=eBay."""
    import csv
    from locg.collection_io import generate_csv
    generate_csv([_make_ready_row()], tmp_path / "out.csv")
    with open(tmp_path / "out.csv", newline="") as f:
        rows = list(csv.reader(f))
    h, d = rows[0], rows[1]
    assert d[h.index("In Collection")] == "1"
    assert d[h.index("In Wish List")] == "0"
    assert d[h.index("Marked Read")] == "0"
    assert d[h.index("Media Format")] == "Print"
    assert d[h.index("Purchase Store")] == "eBay"
    assert d[h.index("Signature")] == "0"
    assert d[h.index("Slabbing")] == "0"


def test_generate_csv_bitforbit_recipe(tmp_path):
    """CSV output for the validated golden fixture rows matches the recipe bit-for-bit."""
    import csv as _csv
    from locg.collection_io import generate_csv

    # Build the same rows as in locg_import_test_recipe.csv
    with open(RECIPE_CSV, newline="") as f:
        reader = _csv.DictReader(f)
        recipe_rows = list(reader)

    # Reconstruct cache rows from the recipe CSV
    cache_rows = []
    for r in recipe_rows:
        cache_rows.append({
            "publisher_name": r["Publisher Name"],
            "series_name": r["Series Name"],
            "full_title": r["Full Title"],
            "release_date": r["Release Date"],
            "price_paid": float(r["Price Paid"]) if r["Price Paid"] else None,
            "date_purchased": r["Date Purchased"] or None,
            "needs_manual_variant": False,
            "needs_manual_series_canonical": False,
            # All other fields not needed for CSV output
        })

    out = tmp_path / "test_out.csv"
    generate_csv(cache_rows, out)

    generated = out.read_text()
    expected = RECIPE_CSV.read_text()
    assert generated == expected


# ---------------------------------------------------------------------------
# Unit 3: generate_notes_md
# ---------------------------------------------------------------------------

def test_notes_md_ready_count(tmp_path):
    """notes.md correctly counts ready rows."""
    from locg.collection_io import generate_notes_md
    ready = [_make_ready_row() for _ in range(5)]
    out = tmp_path / "out.notes.md"
    generate_notes_md(ready, [], [], out)
    text = out.read_text()
    assert "Ready to upload (5 rows)" in text


def test_notes_md_empty_queue(tmp_path):
    """Zero pending rows produces notes.md noting empty queue."""
    from locg.collection_io import generate_notes_md
    out = tmp_path / "out.notes.md"
    generate_notes_md([], [], [], out)
    text = out.read_text()
    assert "Ready to upload (0 rows)" in text


def test_notes_md_manual_variant_section(tmp_path):
    """Variant rows appear in the variants section, not ready section."""
    from locg.collection_io import generate_notes_md
    variant_row = _make_ready_row(full_title="ASM #300 Newsstand")
    variant_row["needs_manual_variant"] = True
    out = tmp_path / "out.notes.md"
    generate_notes_md([], [variant_row], [], out)
    text = out.read_text()
    assert "Needs manual handling — variants (1 rows)" in text
    assert "ASM #300 Newsstand" in text


def test_notes_md_manual_series_section(tmp_path):
    """Series-canonical rows appear in the series canonical section."""
    from locg.collection_io import generate_notes_md
    series_row = _make_ready_row(series="Unknown Series")
    series_row["needs_manual_series_canonical"] = True
    out = tmp_path / "out.notes.md"
    generate_notes_md([], [], [series_row], out)
    text = out.read_text()
    assert "Needs manual handling — series canonical (1 rows)" in text
    assert "Unknown Series" in text


# ---------------------------------------------------------------------------
# Unit 3: _pending_push_rows
# ---------------------------------------------------------------------------

def test_pending_push_rows_partitions(tmp_path):
    """_pending_push_rows correctly partitions ready / manual_variant /
    manual_series / quarantined (the fourth bucket added in BUI-648)."""
    from locg.collection_io import _pending_push_rows

    r = _make_ready_row()
    v = _make_ready_row(full_title="ASM #300 Newsstand")
    v["needs_manual_variant"] = True
    s = _make_ready_row(series="Unknown")
    s["needs_manual_series_canonical"] = True
    already_pushed = _make_ready_row(full_title="Pushed #1")
    already_pushed["pushed_to_locg_at"] = "2030-01-01T00:00:00Z"  # future; not pending
    already_pushed["local_added_at"] = "2026-01-01T00:00:00Z"

    payload = {"comics": [r, v, s, already_pushed]}
    ready, mv, ms, quarantined = _pending_push_rows(payload)
    assert len(ready) == 1
    assert len(mv) == 1
    assert len(ms) == 1
    assert quarantined == []
    assert ready[0]["full_title"] == _make_ready_row()["full_title"]


def test_pending_push_already_pushed_excluded(tmp_path):
    """Rows where local_added_at <= pushed_to_locg_at are not pending."""
    from locg.collection_io import _pending_push_rows

    row = _make_ready_row()
    row["pushed_to_locg_at"] = "2030-01-01T00:00:00Z"  # pushed far in the future
    row["local_added_at"] = "2026-01-01T00:00:00Z"  # added before push timestamp

    ready, mv, ms, quarantined = _pending_push_rows({"comics": [row]})
    assert len(ready) == 0
    assert len(mv) == 0
    assert len(ms) == 0
    assert len(quarantined) == 0


# ---------------------------------------------------------------------------
# import_xlsx — wish-list cache
# ---------------------------------------------------------------------------

def test_import_leaves_existing_wish_list_untouched(tmp_path):
    """BUI-208: import_xlsx no longer touches wish-list.json — an existing cache
    is left byte-for-byte identical (so a server-side removal stays durable)."""
    from locg.collection_io import import_xlsx, wish_list_cache_path

    p = wish_list_cache_path()
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps({
        "updated_at": "2026-05-30T00:00:00+00:00",
        "items": [
            {"name": "Saga #1", "id": None, "source": "local"},
            {"name": "Batman #1", "id": None, "series_name": "Batman (1940 - 2011)",
             "source": "export"},
        ],
    }))
    before = p.read_bytes()

    cache = make_cache(tmp_path)
    import_xlsx(SAMPLE_XLSX, cache)

    assert p.read_bytes() == before, "import must not rewrite wish-list.json"


def test_local_remove_stays_gone_after_import(tmp_path):
    """BUI-206/BUI-208 HEADLINE: a local add removed via cmd_wish_list_remove
    stays gone after a subsequent import — durability is now structural because
    the import never rewrites wish-list.json."""
    from locg.collection_io import import_xlsx, wish_list_cache_path
    import locg.commands as cmds

    cmds.cmd_wish_list_add("Saga #1")
    cmds.cmd_wish_list_add("Daredevil #181")
    cmds.cmd_wish_list_remove("Saga #1")  # the durable removal

    cache = make_cache(tmp_path)
    import_xlsx(SAMPLE_XLSX, cache)

    names = {i["name"] for i in json.loads(wish_list_cache_path().read_text())["items"]}
    assert "Saga #1" not in names, "a removed local wish must not reappear after import"
    assert "Daredevil #181" in names, "the un-removed local wish must remain"


def test_migrate_wish_list_source(tmp_path):
    """BUI-208: migrate stamps an explicit source on every entry, verifies a
    backup, and is idempotent."""
    from locg.collection_io import migrate_wish_list_source, wish_list_cache_path

    p = wish_list_cache_path()
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps({
        "updated_at": "2026-05-30T00:00:00+00:00",
        "items": [
            {"name": "Batman #1", "id": None, "series_name": "Batman (1940 - 2011)"},
            {"name": "Saga #1", "id": None},
        ],
    }))

    result = migrate_wish_list_source()
    assert result["migrated"] == 2
    assert result["total"] == 2

    # Backup exists and verifies byte-for-byte against the pre-migration content.
    backup = Path(result["backup"])
    assert backup.exists()
    assert json.loads(backup.read_text())["items"][1]["name"] == "Saga #1"
    assert "source" not in json.loads(backup.read_text())["items"][1]

    items = {i["name"]: i for i in json.loads(p.read_text())["items"]}
    assert items["Batman #1"]["source"] == "export"
    assert items["Saga #1"]["source"] == "local"

    # Idempotent: a second run stamps nothing new.
    result2 = migrate_wish_list_source()
    assert result2["migrated"] == 0
    assert result2["total"] == 2


def test_migrate_wish_list_source_absent_cache(tmp_path):
    """BUI-208: migrate on a missing cache is a clean no-op."""
    from locg.collection_io import migrate_wish_list_source
    assert migrate_wish_list_source() == {"migrated": 0, "backup": None}


def test_wish_list_add_writes_source_local(tmp_path):
    """BUI-208: cmd_wish_list_add stamps source='local' on the appended entry."""
    import locg.commands as cmds
    from locg.collection_io import wish_list_cache_path

    cmds.cmd_wish_list_add("Saga #1")
    items = json.loads(wish_list_cache_path().read_text())["items"]
    saga = [i for i in items if i["name"] == "Saga #1"]
    assert len(saga) == 1
    assert saga[0]["source"] == "local"


# ---------------------------------------------------------------------------
# wish_rows_for_export (BUI-122): export must never emit In Collection=0 for an
# owned book, and pushes only the local-only (diff) adds, not the full list.
# ---------------------------------------------------------------------------

def _seed_wish(items):
    from locg.collection_io import wish_list_cache_path
    p = wish_list_cache_path()
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps({"updated_at": "2026-01-01T00:00:00+00:00", "items": items}))


def test_wish_export_keeps_local_only_unowned(tmp_path):
    """A local-only add for a book not in the collection is exported."""
    from locg.collection_io import wish_rows_for_export
    _seed_wish([{"name": "Saga #1", "id": None}])
    rows = wish_rows_for_export({"comics": []})
    assert [r["full_title"] for r in rows] == ["Saga #1"]


def test_wish_export_excludes_derived_wishes(tmp_path):
    """A derived wish (carries series_name → LOCG already has it) is NOT exported."""
    from locg.collection_io import wish_rows_for_export
    _seed_wish([{
        "name": "Batman #1", "id": None,
        "series_name": "Batman (1940 - 2011)", "publisher_name": "DC Comics",
        "release_date": "1940-04-25",
    }])
    assert wish_rows_for_export({"comics": []}) == []


def test_wish_export_excludes_owned_book(tmp_path):
    """CRITICAL safety: a local-only add for a book that IS owned is excluded, so
    the CSV can never carry In Collection=0 for it (the deletion bug)."""
    from locg.collection_io import wish_rows_for_export
    _seed_wish([
        {"name": "Marvel Tales #228", "id": None},   # owned -> must be excluded
        {"name": "Hellboy: Wake the Devil #1", "id": None},  # not owned -> kept
    ])
    payload = {"comics": [
        {"full_title": "Marvel Tales #228", "in_collection": 1},
    ]}
    rows = wish_rows_for_export(payload)
    titles = [r["full_title"] for r in rows]
    assert "Marvel Tales #228" not in titles, "owned book must never be a wish row"
    assert "Hellboy: Wake the Devil #1" in titles


def test_wish_export_excludes_owned_xmen_masthead_split(tmp_path):
    """BUI-200 REGRESSION (the 26-deleted-books bug): a wish written under one
    X-Men masthead must NOT be exported as In Collection=0 when the owned copy is
    filed under the OTHER masthead. LOCG files #1-141 under 'The X-Men' and #142+
    under 'Uncanny X-Men'; a literal-title match misses this and the resulting
    In Collection=0 row deletes the owned copy."""
    from locg.collection_io import wish_rows_for_export
    _seed_wish([
        {"name": "Uncanny X-Men #107", "id": None},  # owned as "The X-Men #107"
        {"name": "Uncanny X-Men #142", "id": None},  # genuinely not owned -> kept
    ])
    payload = {"comics": [
        {"full_title": "The X-Men #107", "in_collection": 1},
    ]}
    titles = [r["full_title"] for r in wish_rows_for_export(payload)]
    assert "Uncanny X-Men #107" not in titles, "owned cross-masthead book must never export"
    assert "Uncanny X-Men #142" in titles


def test_wish_export_excludes_owned_when_source_defeats_gate(tmp_path):
    """BUI-208 adversarial: an entry carrying a series_name BUT an explicit
    `source` != 'export' defeats the source gate (it classifies as local and
    reaches the body, unlike the old series_name gate which would have excluded
    it). Deletion safety must then rest on the unconditional owned-safe backstop:
    the owned book is STILL excluded, so no In Collection=0 row is ever emitted
    for it. Locks in the single-layer guarantee the adversarial review relied on."""
    from locg.collection_io import wish_rows_for_export
    _seed_wish([{
        "name": "Uncanny X-Men #107", "id": None,
        "series_name": "Uncanny X-Men",  # would exclude under the OLD gate
        "source": "local",               # contradictory explicit source — defeats the new gate
    }])
    payload = {"comics": [
        {"full_title": "The X-Men #107", "in_collection": 1},  # owned under the other masthead
    ]}
    titles = [r["full_title"] for r in wish_rows_for_export(payload)]
    assert titles == [], "owned book must never export even when the source gate is defeated"


def test_wish_export_excludes_owned_leading_article_variant(tmp_path):
    """BUI-200: owned under a leading-article + decorated series name, wished
    without the article. The normalized (series, issue) match excludes it so no
    In Collection=0 row is emitted for the owned book."""
    from locg.collection_io import wish_rows_for_export
    _seed_wish([{"name": "Incredible Hulk #181", "id": None}])
    payload = {"comics": [
        {"full_title": "The Incredible Hulk #181", "in_collection": 1},
    ]}
    assert wish_rows_for_export(payload) == []


def test_wish_export_owned_match_is_dash_and_article_insensitive(tmp_path):
    """Owned-exclusion normalizes en-dash/hyphen and a leading article, so an
    owned 'Batman: One Bad Day – Two-Face #1' (en-dash) still excludes a wish add
    written with a hyphen / 'The' prefix."""
    from locg.collection_io import wish_rows_for_export
    _seed_wish([{"name": "Batman: One Bad Day - Two-Face #1", "id": None}])  # hyphen
    payload = {"comics": [
        {"full_title": "Batman: One Bad Day – Two-Face #1", "in_collection": 1},  # en-dash
    ]}
    assert wish_rows_for_export(payload) == []


# --- BUI-197: owned-safe export must be masthead-alias aware (delete-prevention) ---

def test_wish_export_excludes_owned_thor_masthead_alias(tmp_path):
    """CRITICAL (BUI-197): a wish written 'The Mighty Thor #300' must NOT be
    exported as In Collection=0 when the owned copy is filed 'Thor #300'. Routing
    the export through the alias-aware owned_match_keys closes the masthead-alias
    variant of the delete bug."""
    from locg.collection_io import wish_rows_for_export
    _seed_wish([
        {"name": "The Mighty Thor #300", "id": None},   # owned as Thor #300
        {"name": "The Mighty Thor #999", "id": None},   # genuinely not owned
    ])
    payload = {"comics": [
        {"full_title": "Thor #300", "in_collection": 1},
    ]}
    titles = [r["full_title"] for r in wish_rows_for_export(payload)]
    assert "The Mighty Thor #300" not in titles, "owned alias book must never export"
    assert "The Mighty Thor #999" in titles


def test_wish_export_excludes_owned_hulk_masthead_alias(tmp_path):
    """BUI-197: wished 'Incredible Hulk #181', owned 'The Incredible Hulk #181'
    (masthead + leading article). Must never emit In Collection=0."""
    from locg.collection_io import wish_rows_for_export
    _seed_wish([{"name": "Incredible Hulk #181", "id": None}])
    payload = {"comics": [
        {"full_title": "The Incredible Hulk #181", "in_collection": 1},
    ]}
    assert wish_rows_for_export(payload) == []


def test_wish_export_excludes_owned_annual_masthead_alias(tmp_path):
    """BUI-197: an annual owned under one masthead, wished under another. Owned
    'Uncanny X-Men Annual #9', wished 'X-Men Annual #9' — must not export."""
    from locg.collection_io import wish_rows_for_export
    _seed_wish([{"name": "X-Men Annual #9", "id": None}])
    payload = {"comics": [
        {"full_title": "Uncanny X-Men Annual #9", "in_collection": 1},
    ]}
    assert wish_rows_for_export(payload) == []


def test_wish_export_excludes_owned_non_digit_issue_token_alias(tmp_path):
    """BUI-197 MUST-FIX 1 (deletion hole): a wish with a NON-digit-led issue token
    ('#A1') owned under an alias name must NOT be emitted In Collection=0. The
    digit-led parser dropped the token, the title-string fallback doesn't alias
    mastheads, so the owned copy would have been deleted. The permissive ownership
    split + alias-aware (series,issue) index now exclude it."""
    from locg.collection_io import wish_rows_for_export
    _seed_wish([{"name": "The Mighty Thor Annual #A1", "id": None}])
    payload = {"comics": [
        {"full_title": "Thor Annual #A1", "in_collection": 1},  # owned under alias name
    ]}
    assert wish_rows_for_export(payload) == [], "owned non-digit-token book must never export"


# --- BUI-197: audit ↔ export parser parity ---

def test_audit_export_parser_parity(tmp_path, monkeypatch):
    """The conflicts audit and the owned-safe export must agree on EVERY wish
    title: an audit conflict ⇒ the export does NOT emit that owned book as
    In Collection=0 (and a non-conflict local-only wish IS emitted). Both now go
    through the single shared split_series_issue_for_ownership parser, so a clean
    audit proves an owned-safe CSV. Crucially this covers NON-digit-led tokens
    (#A1, #annual, #1-A) — the BUI-197 deletion hole, where the digit-led parser
    made such a wish 'unparseable', skipped the ownership check, and exported it
    In Collection=0 over an owned copy filed under an alias name."""
    import locg.commands as cmds

    # A shared owned corpus + a wish set whose tokens stressed the old divergence,
    # including the non-digit-led tokens that reopened the deletion hole.
    owned = [
        ("Thor (Vol. 1) (1966 - 1996)", "Thor #300", "1980-10-01"),
        ("The Incredible Hulk (1968 - 1999)", "The Incredible Hulk #181", "1974-11-01"),
        ("Uncanny X-Men Annual (1980 - 2011)", "Uncanny X-Men Annual #9", "1985-12-01"),
        ("The X-Men (Vol. 1) (1963 - 1981)", "The X-Men #137", "1980-09-01"),
        ("Thor Annual (1966 - 1994)", "Thor Annual #A1", "1966-01-01"),   # non-digit token
        ("The Incredible Hulk (1968 - 1999)", "The Incredible Hulk #annual", "1978-01-01"),
        ("Thor (Vol. 1) (1966 - 1996)", "Thor #1-A", "1966-01-01"),       # hyphen-suffix token
    ]
    wishes = [
        {"name": "The Mighty Thor #300", "id": 1},        # owned via alias
        {"name": "Incredible Hulk #181", "id": 2},        # owned via alias
        {"name": "X-Men Annual #9", "id": 3},             # owned via annual alias
        {"name": "Uncanny X-Men #137", "id": 4},          # owned via split
        {"name": "The Mighty Thor Annual #A1", "id": 7},  # owned via alias + #A1 token
        {"name": "Incredible Hulk #annual", "id": 8},     # owned via alias + word token
        {"name": "The Mighty Thor #1-A", "id": 9},        # owned via alias + #1-A token
        {"name": "The Mighty Thor #999", "id": 5},        # genuinely not owned
        {"name": "Saga #1", "id": 6},                     # genuinely not owned
    ]

    # --- audit side (uses the conftest-isolated cache via CollectionCache) ---
    cache = make_cache(tmp_path)
    monkeypatch.setattr(cmds, "CollectionCache", lambda: cache)
    owned_rows = [
        make_agent_win_row(series=s, full_title=ft, release_date=rd, gixen_item_id=str(i))
        for i, (s, ft, rd) in enumerate(owned)
    ]
    cache.apply(lambda p: p["comics"].extend(owned_rows), command="seed")
    _seed_wish(wishes)  # the audit reads the same wish-list cache the export does
    audit = cmds.cmd_wish_list_conflicts()
    assert audit["unparseable"] == [], "no wish (incl. non-digit tokens) may be skipped"
    conflict_names = {c["name"] for c in audit["conflicts"]}
    assert conflict_names == {
        "The Mighty Thor #300", "Incredible Hulk #181",
        "X-Men Annual #9", "Uncanny X-Men #137",
        "The Mighty Thor Annual #A1", "Incredible Hulk #annual",
        "The Mighty Thor #1-A",
    }

    # --- export side: same wish-list cache (no series_name → local-only adds) and
    # the same owned corpus, so the two paths are compared on identical input.
    from locg.collection_io import wish_rows_for_export
    payload = {"comics": [
        {"full_title": ft, "in_collection": 1} for (_s, ft, _rd) in owned
    ]}
    exported = {r["full_title"] for r in wish_rows_for_export(payload)}

    # PARITY: every audited conflict must be absent from the export (owned-safe).
    for name in conflict_names:
        assert name not in exported, f"audit flagged {name!r} owned but export emitted it"
    # And the genuinely-unowned local-only wishes ARE exported.
    assert "The Mighty Thor #999" in exported
    assert "Saga #1" in exported


# ---------------------------------------------------------------------------
# import_xlsx — BUI-124: hold ownership downgrades in the Phase-2 standard merge.
# The gixen server is the source of truth (BUI-87), so a LOCG export reporting
# In Collection=0 over an owned row must NOT silently un-own the book (which
# would make collection-check buy a duplicate). The downgrade is held (existing
# in_collection preserved), flagged as ownership_downgrade_held, and counted.
# ---------------------------------------------------------------------------

def _import_owned_then(tmp_path: Path, first: dict[str, Any], second: dict[str, Any]):
    """Import `first` (establishes an owned locg_export row), then `second`
    (the candidate downgrade). Returns (cache, second_result)."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    xlsx1 = tmp_path / "first.xlsx"
    _build_export_xlsx(xlsx1, [first])
    import_xlsx(xlsx1, cache)

    xlsx2 = tmp_path / "second.xlsx"
    _build_export_xlsx(xlsx2, [second])
    result = import_xlsx(xlsx2, cache)
    return cache, result


def _audit_types(tmp_path: Path) -> list[str]:
    path = tmp_path / "import-history.jsonl"
    if not path.exists():
        return []
    lines = path.read_text().strip().splitlines()
    return [json.loads(l)["type"] for l in lines if l]


def test_exact_identity_downgrade_held(tmp_path):
    """Exact-identity update: an owned row stays owned when the re-export says
    In Collection=0; the hold is flagged and counted."""
    base = {
        "publisher": "Marvel Comics", "series": "Daredevil",
        "full_title": "Daredevil #181", "release_date": "1982-04-10",
    }
    cache, result = _import_owned_then(
        tmp_path,
        {**base, "in_collection": 1},
        {**base, "in_collection": 0},
    )

    rows = [r for r in cache.load()["comics"] if r["full_title"] == "Daredevil #181"]
    assert len(rows) == 1
    assert rows[0]["in_collection"] == 1, "ownership must be preserved, not downgraded"
    assert result["ownership_downgrades_held"] == 1
    assert "ownership_downgrade_held" in _audit_types(tmp_path)


def test_rename_branch_downgrade_held(tmp_path):
    """Rename branch (same publisher/series/release_date, new full_title): an
    owned row stays owned when the renamed export row says In Collection=0."""
    cache, result = _import_owned_then(
        tmp_path,
        {"publisher": "Marvel Comics", "series": "Amazing Spider-Man",
         "full_title": "Amazing Spider-Man #300", "release_date": "1988-05-10",
         "in_collection": 1},
        {"publisher": "Marvel Comics", "series": "Amazing Spider-Man",
         "full_title": "Amazing Spider-Man #300 Direct", "release_date": "1988-05-10",
         "in_collection": 0},
    )

    rows = cache.load()["comics"]
    renamed = [r for r in rows if r["full_title"] == "Amazing Spider-Man #300 Direct"]
    assert len(renamed) == 1, "rename must update in place, not insert"
    assert renamed[0]["in_collection"] == 1, "ownership preserved across the rename"
    assert renamed[0]["previous_full_title"] == "Amazing Spider-Man #300"
    assert result["ownership_downgrades_held"] == 1
    assert "ownership_downgrade_held" in _audit_types(tmp_path)


def test_non_downgrade_copies_in_collection_normally(tmp_path):
    """A re-export that keeps the book owned (In Collection=1) copies through and
    is NOT flagged as a held downgrade."""
    base = {
        "publisher": "Marvel Comics", "series": "Daredevil",
        "full_title": "Daredevil #181", "release_date": "1982-04-10",
    }
    cache, result = _import_owned_then(
        tmp_path,
        {**base, "in_collection": 1},
        {**base, "in_collection": 1},
    )

    rows = [r for r in cache.load()["comics"] if r["full_title"] == "Daredevil #181"]
    assert rows[0]["in_collection"] == 1
    assert result["ownership_downgrades_held"] == 0
    assert "ownership_downgrade_held" not in _audit_types(tmp_path)


def test_count_decrease_that_stays_owned_copies_normally(tmp_path):
    """in_collection is a copies-owned count; a decrease that stays truthy
    (2 -> 1) is a normal update, not a downgrade — it copies through unflagged."""
    base = {
        "publisher": "Marvel Comics", "series": "Daredevil",
        "full_title": "Daredevil #181", "release_date": "1982-04-10",
    }
    cache, result = _import_owned_then(
        tmp_path,
        {**base, "in_collection": 2},
        {**base, "in_collection": 1},
    )

    rows = [r for r in cache.load()["comics"] if r["full_title"] == "Daredevil #181"]
    assert rows[0]["in_collection"] == 1, "count change that stays owned applies"
    assert result["ownership_downgrades_held"] == 0
    assert "ownership_downgrade_held" not in _audit_types(tmp_path)


# ---------------------------------------------------------------------------
# import_xlsx — BUI-412: non-blocking data-quality report for owned rows with
# no release_date. A null-dated owned row silently defeats the year-scoped
# wish-list conflicts audit (the year-gate can't confirm two years differ
# against a null-dated owned row). The importer must surface — never guard
# against — this gap: count it, warn about it, and store the row unchanged.
# ---------------------------------------------------------------------------

def test_import_reports_null_release_date_on_owned_row(tmp_path):
    """An owned row with an empty release_date is counted + warned about, and
    is still stored unchanged (non-blocking: no reject, no drop, no alter)."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    xlsx = tmp_path / "import.xlsx"
    _build_export_xlsx(xlsx, [
        {
            "publisher": "Marvel Comics", "series": "Fantastic Four (Vol. 7)",
            "full_title": "Fantastic Four (Vol. 7) #4", "release_date": "",
            "in_collection": 1, "in_wish_list": 0,
        },
    ])

    result = import_xlsx(xlsx, cache)

    # The report fired.
    assert result["null_release_date_owned"] == 1
    assert any("release_date" in w and "BUI-412" in w for w in result["warnings"]), (
        f"expected a BUI-412 release_date warning, got: {result['warnings']}"
    )

    # The row is still present, unchanged (added, not rejected/dropped/altered).
    assert result["added"] == 1
    rows = [
        r for r in cache.load()["comics"]
        if r["full_title"] == "Fantastic Four (Vol. 7) #4"
    ]
    assert len(rows) == 1, "row must still be stored"
    assert rows[0]["in_collection"] == 1
    assert not (rows[0]["release_date"] or ""), "release_date must be left null/empty, not backfilled"


def test_import_null_release_date_report_excludes_dated_and_wish_only_rows(tmp_path):
    """The count/warning must fire only for OWNED rows with a null/empty
    release_date — not for a dated owned row, and not for a wish-only row
    (in_collection=0) that also happens to lack a release_date."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    xlsx = tmp_path / "import.xlsx"
    _build_export_xlsx(xlsx, [
        {
            "publisher": "Marvel Comics", "series": "Daredevil",
            "full_title": "Daredevil #181", "release_date": "1982-04-10",
            "in_collection": 1, "in_wish_list": 0,
        },
        {
            "publisher": "Marvel Comics", "series": "Saga",
            "full_title": "Saga #1", "release_date": "",
            "in_collection": 0, "in_wish_list": 1,
        },
    ])

    result = import_xlsx(xlsx, cache)

    assert result["null_release_date_owned"] == 0
    assert not any("BUI-412" in w for w in result["warnings"])


# ---------------------------------------------------------------------------
# BUI-548: the reconciler must survive the drift between what record-win writes
# and what LOCG hands back — provider publisher vocabulary, cover-vs-on-sale
# release dates, punctuation/article/whitespace in the name — without ever
# merging two genuinely different books.
#
# Root cause of the 2026-07-27 incident (41 pending wins, only 13 matched):
# BUI-458 started stamping Metron's publisher label on agent_win rows, and
# `_publisher_matches` compared it verbatim against LOCG's. "Marvel" vs "Marvel
# Comics" scored 0 and blocked 34 of the 41 outright.
# ---------------------------------------------------------------------------

def test_normalize_publisher_folds_provider_vocabulary():
    """The same company named by Metron and by LOCG lands on one key."""
    from locg.collection_io import _normalize_publisher, _publisher_matches

    assert _normalize_publisher("Marvel") == _normalize_publisher("Marvel Comics")
    assert _normalize_publisher("DC Comics") == "dc"
    assert _normalize_publisher("Boom! Studios") == _normalize_publisher("BOOM! Studios")
    assert _normalize_publisher("Dark Horse Comics") == "dark horse"
    # A name that is ENTIRELY generic words must not fold to nothing.
    assert _normalize_publisher("Comics") == "comics"
    # Genuinely different publishers stay apart.
    assert not _publisher_matches("Marvel", "DC Comics")
    assert not _publisher_matches("Skybound", "Image Comics")
    # A missing side is still a wildcard.
    assert _publisher_matches(None, "Marvel Comics")


def test_pending_win_reconciles_across_metron_vs_locg_publisher(tmp_path):
    """BUI-548 regression, `The X-Men #14` from the 2026-07-27 sync: publisher
    'Marvel' (Metron, via BUI-458) vs 'Marvel Comics' (LOCG), plus the
    cover-date/on-sale skew 1965-11-01 -> 1965-09-02."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    win = make_agent_win_row(
        publisher="Marvel",  # what record-win stamps since BUI-458
        series="The X-Men (Vol. 1) (1963 - 1981)",
        full_title="The X-Men #14",
        release_date="1965-11-01",  # Metron cover date
        needs_manual_series=False,
        pushed=None,
    )

    def add_win(payload):
        payload["comics"].append(win)

    cache.apply(add_win, command="pre-import")

    xlsx = tmp_path / "reexport.xlsx"
    _build_export_xlsx(xlsx, [{
        "publisher": "Marvel Comics",
        "series": "The X-Men (Vol. 1) (1963 - 1981)",
        "full_title": "The X-Men #14",
        "release_date": "1965-09-02",  # LOCG on-sale date
    }])

    result = import_xlsx(xlsx, cache)
    payload = cache.load()

    rows = [r for r in payload["comics"] if r["full_title"] == "The X-Men #14"]
    assert len(rows) == 1, "must reconcile in place, not create a second owned row"
    assert rows[0]["pushed_to_locg_at"] is not None
    assert rows[0]["publisher_name"] == "Marvel Comics"
    assert result["reconciled"] == 1
    assert result["added"] == 0
    assert result["owned_duplicate_identities"] == 0


def test_cover_date_skew_across_a_year_boundary_reconciles(tmp_path):
    """BUI-548, `Tales of Suspense #98`: cover 1968-02-01 vs on-sale 1967-11-02.
    Same issue, 91 days apart, different years — the old exact-year gate scored
    it 0 and duplicated it."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    win = make_agent_win_row(
        publisher="Marvel",
        series="Tales of Suspense (Vol. 1) (1958 - 1967)",
        full_title="Tales of Suspense #98",
        release_date="1968-02-01",
        needs_manual_series=False,
        pushed=None,
    )

    def add_win(payload):
        payload["comics"].append(win)

    cache.apply(add_win, command="pre-import")

    xlsx = tmp_path / "reexport.xlsx"
    _build_export_xlsx(xlsx, [{
        "publisher": "Marvel Comics",
        "series": "Tales of Suspense (Vol. 1) (1958 - 1967)",
        "full_title": "Tales of Suspense #98",
        "release_date": "1967-11-02",
    }])

    result = import_xlsx(xlsx, cache)
    payload = cache.load()

    rows = [r for r in payload["comics"] if r["full_title"] == "Tales of Suspense #98"]
    assert len(rows) == 1
    assert rows[0]["release_date"] == "1967-11-02"
    assert result["reconciled"] == 1
    assert result["added"] == 0


def test_date_window_is_asymmetric_and_bounded():
    """The cover-vs-on-sale window is one-directional (LOCG earlier) and short.

    An export date LATER than the win's is not this skew and gets no tolerance
    at all; a gap beyond the window is a different printing/volume, not a
    re-dating."""
    from locg.collection_io import _release_dates_compatible

    def pair(win_date, export_date):
        return _release_dates_compatible(
            {"release_date": win_date}, {"release_date": export_date}
        )

    assert pair("1968-02-01", "1967-11-02"), "91 days earlier — the real skew"
    assert not pair("1968-02-01", "1967-06-01"), "245 days is out of window"
    assert not pair("1967-11-02", "1968-02-01"), "export LATER is never this skew"
    assert pair("1965-11-01", "1965-09-02"), "same year still passes, as before"
    assert pair("1987-01-01", None), "fails open when a side has no date"


def test_reconcile_score_never_merges_a_base_issue_into_a_printing():
    """BUI-548 hard guard, at the exact gate that enforces it.

    `Batman: The Dark Knight Returns #2` and its `#2 3rd Printing` are distinct
    collectibles and were both in the store at once. LOCG spells an edition as
    text trailing the issue number, which the end-anchored token extractor reads
    as "no token" — so a base win scores 0 against it. Asserted with the
    STRONGEST corroboration the reconciler has (identical price, purchase date
    AND release date) to prove no widening can reach past it.

    Passes before BUI-548 too, by design: it pins behavior the widening must
    NOT change, not behavior the widening adds."""
    from locg.collection_io import _reconcile_score

    base = {
        "publisher_name": "DC Comics",
        "series_name": "Batman: The Dark Knight Returns (1986)",
        "full_title": "Batman: The Dark Knight Returns #2",
        "release_date": "1986-03-31",
        "price_paid": 22.5,
        "date_purchased": "2026-06-05",
    }
    third_printing = dict(
        base, full_title="Batman: The Dark Knight Returns #2 3rd Printing"
    )
    assert _reconcile_score(base, third_printing) == 0
    # ...and a newsstand/direct edition, spelled the same way.
    newsstand = {
        "publisher_name": "Marvel",
        "series_name": "The Avengers (Vol. 1) (1963 - 1996)",
        "full_title": "The Avengers #196 Newsstand Edition",
        "release_date": "1980-03-18",
        "price_paid": 12.0,
        "date_purchased": "2026-07-01",
    }
    base_avengers = dict(newsstand, full_title="The Avengers #196")
    assert _reconcile_score(base_avengers, newsstand) == 0


def test_different_printing_of_the_same_issue_never_merges(tmp_path):
    """BUI-548 hard guard, end to end. Production dates: a reprint ships on its
    own on-sale date (base 1986-03-25, 3rd printing 1986-03-31), and both rows
    must survive the import as two books."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    win = make_agent_win_row(
        publisher="DC Comics",
        series="Batman: The Dark Knight Returns (1986)",
        full_title="Batman: The Dark Knight Returns #2",
        release_date="1986-03-25",
        needs_manual_series=False,
        pushed=None,
    )
    # Same purchase fingerprint on both sides — the strongest corroboration
    # signal the reconciler has. It still must not merge two printings.
    win["price_paid"] = 22.5
    win["date_purchased"] = "2026-06-05"

    def add_win(payload):
        payload["comics"].append(win)

    cache.apply(add_win, command="pre-import")

    xlsx = tmp_path / "reexport.xlsx"
    _build_export_xlsx(xlsx, [{
        "publisher": "DC Comics",
        "series": "Batman: The Dark Knight Returns (1986)",
        "full_title": "Batman: The Dark Knight Returns #2 3rd Printing",
        "release_date": "1986-03-31",
        "price_paid": 22.5,
        "date_purchased": "2026-06-05",
    }])

    result = import_xlsx(xlsx, cache)
    payload = cache.load()

    titles = sorted(
        r["full_title"] for r in payload["comics"]
        if r["full_title"].startswith("Batman: The Dark Knight Returns #2")
    )
    assert titles == [
        "Batman: The Dark Knight Returns #2",
        "Batman: The Dark Knight Returns #2 3rd Printing",
    ], "the two printings must remain two rows"
    assert result["reconciled"] == 0
    assert result["auto_healed_duplicates"] == 0
    assert result["owned_duplicate_identities"] == 0, (
        "a base issue and its 3rd printing are two books, not a duplicate pair"
    )


def test_two_declared_volumes_that_disagree_never_merge():
    """BUI-548: corroboration widens a MISSING volume annotation, never a
    contradicting one. `Silver Surfer (Vol. 3)` vs `The Silver Surfer (Vol. 4)`
    is positive evidence of two books; it stays a hard -1 even with an identical
    price + purchase date.

    Passes before BUI-548 too, by design: it pins behavior the widening must
    NOT change, not behavior the widening adds."""
    from locg.collection_io import _reconcile_score

    win = {
        "publisher_name": "Marvel",
        "series_name": "Silver Surfer (Vol. 3) (1987 - 1998)",
        "full_title": "Silver Surfer #1",
        "release_date": "1988-12-01",
        "price_paid": 16.5,
        "date_purchased": "2026-07-18",
    }
    export = {
        "publisher_name": "Marvel Comics",
        "series_name": "The Silver Surfer (Vol. 4) (1988)",
        "full_title": "The Silver Surfer #1",
        "release_date": "1988-12-01",
        "price_paid": 16.5,
        "date_purchased": "2026-07-18",
    }
    assert _reconcile_score(win, export) == -1


def test_whitespace_only_series_drift_merges_only_with_corroboration():
    """BUI-548 / BUI-546: `Dawn Runner` (eBay) vs `Dawnrunner` (LOCG) produced 5
    duplicate owned rows. BUI-546 deliberately did NOT widen
    `_normalize_series_key` to strip inner whitespace — that key also drives the
    ownership matcher. So the squash lives in the reconciler and fires only when
    an independent signal agrees; here the release dates match exactly."""
    from locg.collection_io import _reconcile_score

    win = {
        "publisher_name": "Dark Horse Comics",
        "series_name": "Dawn Runner",
        "full_title": "Dawn Runner #1",
        "release_date": "2024-03-20",
    }
    export = {
        "publisher_name": "Dark Horse Comics",
        "series_name": "Dawnrunner (2024)",
        "full_title": "Dawnrunner #1",
        "release_date": "2024-03-20",
    }
    assert _reconcile_score(win, export) > 0

    # Same names, but nothing corroborates: different release dates, no shared
    # purchase fingerprint. The whitespace squash must NOT fire on its own.
    uncorroborated = dict(export, release_date="2024-06-20")
    assert _reconcile_score(win, uncorroborated) == 0


def test_punctuation_class_auto_heals_against_an_owned_row(tmp_path):
    """BUI-548 + BUI-546, the `Doctor Strange, Sorcerer Supreme` class: the win
    drops the comma, LOCG keeps it. With punctuation folded into the series key
    the win now reaches the collision guard, and — the book being already
    owned, same year, same base edition — BUI-211's auto-heal retires it."""
    from locg.collection_io import import_xlsx, make_identity

    cache = make_cache(tmp_path)
    owned = make_agent_win_row(
        publisher="Marvel Comics",
        series="Doctor Strange, Sorcerer Supreme (1988 - 1996)",
        full_title="Doctor Strange, Sorcerer Supreme #44",
        release_date="1992-06-16",
        gixen_item_id=None,
        pushed="2024-01-01T00:00:00.000000Z",
    )
    owned["source"] = "locg_export"
    win = make_agent_win_row(
        publisher="Marvel",
        series="Doctor Strange Sorcerer Supreme",  # comma dropped
        full_title="Doctor Strange Sorcerer Supreme #44",
        release_date="1992-08-01",  # cover date
        gixen_item_id="99",
        needs_manual_series=False,
        pushed=None,
    )

    def add_rows(payload):
        payload["comics"].extend([owned, win])

    cache.apply(add_rows, command="pre-import")

    xlsx = tmp_path / "reexport.xlsx"
    _build_export_xlsx(xlsx, [{
        "publisher": "Marvel Comics",
        "series": "Doctor Strange, Sorcerer Supreme (1988 - 1996)",
        "full_title": "Doctor Strange, Sorcerer Supreme #44",
        "release_date": "1992-06-16",
    }])

    result = import_xlsx(xlsx, cache)
    payload = cache.load()

    rows = [r for r in payload["comics"] if "Sorcerer Supreme #44" in r["full_title"]]
    assert len(rows) == 1, "the duplicate pair must collapse to one row"
    assert rows[0]["source"] == "locg_export"
    assert rows[0]["gixen_item_id"] == "99", "the win's bid link is carried over"
    assert result["auto_healed_duplicates"] == 1
    assert result["owned_duplicate_identities"] == 0
    idents = [make_identity(r) for r in payload["comics"]]
    assert len(idents) == len(set(idents))


def test_widening_never_reaches_the_delete_path_across_a_year_boundary(tmp_path):
    """BUI-548 blast-radius bound: the date window widens which rows the
    reconciler MATCHES, but `_era_confirmed` (BUI-462) still demands an exact
    same-year match before anything is deleted. A cross-year match that collides
    with an owned row is left pending with a named warning — visible non-clear
    over silent wrong drop."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    owned = make_agent_win_row(
        publisher="Marvel Comics",
        series="Fantastic Four (Vol. 1) (1961 - 1996)",
        full_title="Fantastic Four #46",
        release_date="1965-10-12",
        gixen_item_id=None,
        pushed="2024-01-01T00:00:00.000000Z",
    )
    owned["source"] = "locg_export"
    win = make_agent_win_row(
        publisher="Marvel",
        series="Fantastic Four (Vol. 1) (1961 - 1996)",
        full_title="Fantastic Four #46",
        release_date="1966-01-01",  # 81 days later, different year
        gixen_item_id="99",
        needs_manual_series=False,
        pushed=None,
    )

    def add_rows(payload):
        payload["comics"].extend([owned, win])

    cache.apply(add_rows, command="pre-import")

    xlsx = tmp_path / "reexport.xlsx"
    _build_export_xlsx(xlsx, [{
        "publisher": "Marvel Comics",
        "series": "Fantastic Four (Vol. 1) (1961 - 1996)",
        "full_title": "Fantastic Four #46",
        "release_date": "1965-10-12",
    }])

    result = import_xlsx(xlsx, cache)
    payload = cache.load()

    rows = [r for r in payload["comics"] if r["full_title"] == "Fantastic Four #46"]
    assert len(rows) == 2, "nothing may be deleted on cross-year evidence"
    assert result["auto_healed_duplicates"] == 0
    assert any("release years disagree" in w for w in result["warnings"])
    # ...and the semantic duplicate check names it rather than letting the sync
    # report clean.
    assert result["owned_duplicate_identities"] == 1


# ---------------------------------------------------------------------------
# BUI-548: post-import semantic duplicate check. The sync's row-count
# arithmetic balanced EXACTLY on 2026-07-27 while 28 books quietly became owned
# twice, because every duplicate is one `added` row.
# ---------------------------------------------------------------------------

def test_owned_duplicate_identities_flags_an_unreconciled_twin(tmp_path):
    """A pending win and an export row claiming the same book, left unmatched,
    must be counted and named — not absorbed into `added`."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    win = make_agent_win_row(
        publisher="Marvel",
        # A DECLARED volume conflict — the one thing corroboration never widens,
        # so this row genuinely cannot reconcile.
        series="Strange Tales (Vol. 2) (1987 - 1988)",
        full_title="Strange Tales #135",
        release_date="1965-08-01",
        needs_manual_series=False,
        pushed=None,
    )

    def add_win(payload):
        payload["comics"].append(win)

    cache.apply(add_win, command="pre-import")

    xlsx = tmp_path / "reexport.xlsx"
    _build_export_xlsx(xlsx, [{
        "publisher": "Marvel Comics",
        "series": "Strange Tales (Vol. 1) (1951 - 1976)",
        "full_title": "Strange Tales #135",
        "release_date": "1965-05-04",
    }])

    result = import_xlsx(xlsx, cache)

    assert result["added"] == 1, "the arithmetic sees only ordinary growth"
    assert result["owned_duplicate_identities"] == 1
    assert any("owned TWICE" in w for w in result["warnings"])
    assert any("strangetales#135" in w for w in result["warnings"])


def test_owned_duplicate_check_ignores_two_genuine_volumes(tmp_path):
    """`X-Men #128` (Vol. 2, 2002) and `The X-Men #128` (Vol. 1, 1979) are two
    books legitimately owned side by side. A title-only check would cry
    duplicate on them and train the operator to ignore it, so the pair must also
    be date-compatible before it counts."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    win = make_agent_win_row(
        publisher=None,
        series="X-Men (Vol. 2) (2001 - 2013)",
        full_title="X-Men #128",
        release_date="2002-01-01",
        needs_manual_series=False,
        pushed=None,
    )

    def add_win(payload):
        payload["comics"].append(win)

    cache.apply(add_win, command="pre-import")

    xlsx = tmp_path / "reexport.xlsx"
    _build_export_xlsx(xlsx, [{
        "publisher": "Marvel Comics",
        "series": "The X-Men (Vol. 1) (1963 - 1981)",
        "full_title": "The X-Men #128",
        "release_date": "1979-09-18",
    }])

    result = import_xlsx(xlsx, cache)

    assert result["reconciled"] == 0, "different volumes must not merge"
    assert result["owned_duplicate_identities"] == 0, "and must not be reported as dupes"


# ---------------------------------------------------------------------------
# BUI-547: `needs_manual_series_canonical` is written once, at record-win time,
# and nothing ever recomputes it — so rows enter the manual bucket and never
# leave it, even after the collection gains the data that would resolve them.
# ---------------------------------------------------------------------------

def test_stale_manual_series_flag_clears_and_row_reconciles(tmp_path):
    """BUI-547 acceptance, `Infinity Gauntlet #2`: flagged on 2026-07-02 when the
    index had no entry for the key. `The Infinity Gauntlet #1` has since arrived
    via import, so the key resolves — the flag must clear, the canonical name
    must be written, and (BUI-548) the row must then reconcile in the SAME
    import rather than exporting one more duplicate first."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    # An already-imported sibling issue — this is what puts the series in the index.
    sibling = make_agent_win_row(
        publisher="Marvel Comics",
        series="The Infinity Gauntlet (Vol. 1) (1991)",
        full_title="The Infinity Gauntlet #1",
        release_date="1991-05-21",
        gixen_item_id=None,
        pushed="2024-01-01T00:00:00.000000Z",
    )
    sibling["source"] = "locg_export"
    flagged = make_agent_win_row(
        publisher="Marvel",
        series="Infinity Gauntlet",          # bare Metron masthead
        full_title="Infinity Gauntlet #2",   # no leading article
        release_date="1991-08-01",           # cover date
        gixen_item_id="99",
        needs_manual_series=True,
        pushed=None,
    )

    def add_rows(payload):
        payload["comics"].extend([sibling, flagged])

    cache.apply(add_rows, command="pre-import")

    xlsx = tmp_path / "reexport.xlsx"
    _build_export_xlsx(xlsx, [
        {"publisher": "Marvel Comics", "series": "The Infinity Gauntlet (Vol. 1) (1991)",
         "full_title": "The Infinity Gauntlet #1", "release_date": "1991-05-21"},
        {"publisher": "Marvel Comics", "series": "The Infinity Gauntlet (Vol. 1) (1991)",
         "full_title": "The Infinity Gauntlet #2", "release_date": "1991-06-18"},
    ])

    result = import_xlsx(xlsx, cache)
    payload = cache.load()

    assert result["manual_series_flags_cleared"] == 1
    issue2 = [r for r in payload["comics"] if r["full_title"] == "The Infinity Gauntlet #2"]
    assert len(issue2) == 1, "must reconcile, not insert a second owned row"
    assert issue2[0]["needs_manual_series_canonical"] is False
    assert issue2[0]["series_name"] == "The Infinity Gauntlet (Vol. 1) (1991)"
    assert issue2[0]["gixen_item_id"] == "99", "it is the same row, re-identified"
    assert result["owned_duplicate_identities"] == 0
    # A wrong clear must be reversible from the append-only log alone, so the
    # record carries the name it replaced.
    audit = [
        json.loads(line)
        for line in (tmp_path / "import-history.jsonl").read_text().splitlines()
    ]
    cleared = [r for r in audit if r["type"] == "manual_series_flag_cleared"]
    assert len(cleared) == 1
    assert cleared[0]["details"]["previous_series_name"] == "Infinity Gauntlet"
    assert cleared[0]["details"]["resolved_series_name"] == (
        "The Infinity Gauntlet (Vol. 1) (1991)"
    )


def test_manual_series_flag_clears_for_a_series_arriving_in_this_export(tmp_path):
    """BUI-547: the operator adds the missing series to LOCG precisely to unstick
    a flagged win. `import_xlsx` rebuilds `series_name_index` at the very END, so
    resolving against the STORED index would still miss — the manual work would
    not have worked. The pass resolves against the index as it will exist after
    this import, which includes the incoming export."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    flagged = make_agent_win_row(
        publisher="BOOM! Studios",
        series="Rare Flavours",
        full_title="Rare Flavours #7",
        release_date="2024-08-28",
        gixen_item_id="99",
        needs_manual_series=True,
        pushed=None,
    )

    def add_row(payload):
        payload["comics"].append(flagged)

    cache.apply(add_row, command="pre-import")

    xlsx = tmp_path / "reexport.xlsx"
    # The series arrives NOW, via a different issue. Nothing in the store knew it.
    _build_export_xlsx(xlsx, [{
        "publisher": "BOOM! Studios", "series": "Rare Flavours (2023 - 2024)",
        "full_title": "Rare Flavours #1", "release_date": "2023-09-20",
    }])

    result = import_xlsx(xlsx, cache)
    payload = cache.load()

    row = next(r for r in payload["comics"] if r["gixen_item_id"] == "99")
    assert result["manual_series_flags_cleared"] == 1
    assert row["needs_manual_series_canonical"] is False
    assert row["series_name"] == "Rare Flavours (2023 - 2024)"
    # Still pending (LOCG has no #7 yet) — but no longer withheld from the CSV.
    assert row["pushed_to_locg_at"] is None


def test_unresolvable_row_keeps_its_manual_series_flag(tmp_path):
    """BUI-547 negative: a row whose series still doesn't resolve keeps its flag
    and stays out of the CSV. No false clearing."""
    from locg.collection_io import _pending_push_rows, import_xlsx

    cache = make_cache(tmp_path)
    flagged = make_agent_win_row(
        publisher="Marvel",
        series="Some Series Nobody Has Ever Heard Of",
        full_title="Some Series Nobody Has Ever Heard Of #1",
        release_date="1999-01-01",
        gixen_item_id="99",
        needs_manual_series=True,
        pushed=None,
    )

    def add_row(payload):
        payload["comics"].append(flagged)

    cache.apply(add_row, command="pre-import")

    xlsx = tmp_path / "reexport.xlsx"
    _build_export_xlsx(xlsx, [{
        "publisher": "Marvel Comics", "series": "Daredevil (Vol. 1) (1964 - 1998)",
        "full_title": "Daredevil #181", "release_date": "1982-04-10",
    }])

    result = import_xlsx(xlsx, cache)
    payload = cache.load()

    row = next(r for r in payload["comics"] if r["gixen_item_id"] == "99")
    assert result["manual_series_flags_cleared"] == 0
    assert row["needs_manual_series_canonical"] is True
    assert row["series_name"] == "Some Series Nobody Has Ever Heard Of"
    _ready, _variant, manual_series, _quarantined = _pending_push_rows(payload)
    assert [r["gixen_item_id"] for r in manual_series] == ["99"], "stays out of the CSV"


def test_manual_series_pass_never_sets_the_flag(tmp_path):
    """BUI-547 is one-way on purpose. An UNflagged row whose series no longer
    resolves must not be silently re-flagged out of the export — that is a
    different failure with a different blast radius."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    unflagged = make_agent_win_row(
        publisher="Marvel",
        series="A Series The Export No Longer Carries",
        full_title="A Series The Export No Longer Carries #3",
        release_date="1999-01-01",
        gixen_item_id="99",
        needs_manual_series=False,
        pushed=None,
    )

    def add_row(payload):
        payload["comics"].append(unflagged)

    cache.apply(add_row, command="pre-import")

    xlsx = tmp_path / "reexport.xlsx"
    _build_export_xlsx(xlsx, [{
        "publisher": "Marvel Comics", "series": "Daredevil (Vol. 1) (1964 - 1998)",
        "full_title": "Daredevil #181", "release_date": "1982-04-10",
    }])

    import_xlsx(xlsx, cache)
    payload = cache.load()

    row = next(r for r in payload["comics"] if r["gixen_item_id"] == "99")
    assert row["needs_manual_series_canonical"] is False, "never re-flagged"


def test_flag_clear_rekeys_the_identity_indices(tmp_path):
    """BUI-547 index hygiene: clearing a flag rewrites `series_name`, which is
    part of BOTH index keys. A stale entry would let an unrelated export row
    claim the row's slot as a rename target and apply its columns to it — the
    hazard BUI-462 fixed for the auto-heal drop. Here the export carries a row
    whose partial identity equals the flagged row's PRE-clear one; it must
    insert cleanly instead of hijacking the win."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    seed = make_agent_win_row(
        publisher="Marvel Comics",
        series="Nova (Vol. 1) (1976 - 1979)",
        full_title="Nova #1",
        release_date="1976-06-01",
        gixen_item_id=None,
        pushed="2024-01-01T00:00:00.000000Z",
    )
    seed["source"] = "locg_export"
    flagged = make_agent_win_row(
        publisher="Marvel Comics",
        series="Nova",  # unresolved bare masthead — the PRE-clear series_name
        full_title="Nova #99",
        release_date="1976-06-01",  # same date as the decoy below
        gixen_item_id="99",
        needs_manual_series=True,
        pushed=None,
    )

    def add_rows(payload):
        payload["comics"].extend([seed, flagged])

    cache.apply(add_rows, command="pre-import")

    xlsx = tmp_path / "reexport.xlsx"
    _build_export_xlsx(xlsx, [
        {"publisher": "Marvel Comics", "series": "Nova (Vol. 1) (1976 - 1979)",
         "full_title": "Nova #1", "release_date": "1976-06-01"},
        # Partial identity ("Marvel Comics", "Nova", "1976-06-01") — exactly the
        # flagged row's key BEFORE the clear.
        {"publisher": "Marvel Comics", "series": "Nova",
         "full_title": "Nova #2", "release_date": "1976-06-01"},
    ])

    import_xlsx(xlsx, cache)
    payload = cache.load()

    win = next(r for r in payload["comics"] if r["gixen_item_id"] == "99")
    assert win["series_name"] == "Nova (Vol. 1) (1976 - 1979)", "flag cleared + renamed"
    assert win["full_title"] == "Nova #99", (
        "the stale index slot must not let an unrelated export row rename this win"
    )
    assert win["previous_full_title"] is None


# ---------------------------------------------------------------------------
# BUI-586: publisher-scope the volume pool `_reresolve_manual_series_flags`
# draws on, so a re-resolution can't land on a same-named volume from the
# wrong publisher (BUI-564's trigger-and-rescope pattern, minus the Metron
# fetch — the row's own `publisher_name` is already on hand).
# ---------------------------------------------------------------------------

def test_reresolve_manual_series_no_conflict_keeps_unscoped_answer_bui586(tmp_path):
    """Both known volumes under the key agree with the win's publisher, so
    `series_publisher_conflicts` never fires and the rescope branch is never
    entered — the outcome must be byte-identical to the pre-BUI-586 plain
    era lookup (narrowest range wins)."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    older = make_agent_win_row(
        publisher="Image Comics",
        series="Spawn (1992 - Present)",
        full_title="Spawn #1",
        release_date="1992-05-01",
        gixen_item_id=None,
        pushed="2024-01-01T00:00:00.000000Z",
    )
    older["source"] = "locg_export"
    newer = make_agent_win_row(
        publisher="Image Comics",
        series="Spawn (2012 - Present)",
        full_title="Spawn #224",
        release_date="2012-10-01",
        gixen_item_id=None,
        pushed="2024-01-01T00:00:00.000000Z",
    )
    newer["source"] = "locg_export"
    flagged = make_agent_win_row(
        publisher="Image Comics",
        series="Spawn",  # bare masthead — unresolved
        full_title="Spawn #250",
        release_date="2015-01-01",
        gixen_item_id="99",
        needs_manual_series=True,
        pushed=None,
    )

    def add_rows(payload):
        payload["comics"].extend([older, newer, flagged])

    cache.apply(add_rows, command="pre-import")

    xlsx = tmp_path / "reexport.xlsx"
    _build_export_xlsx(xlsx, [{
        "publisher": "DC Comics", "series": "Batman (Vol. 3) (2016 - Present)",
        "full_title": "Batman #1", "release_date": "2016-06-15",
    }])

    result = import_xlsx(xlsx, cache)
    payload = cache.load()

    row = next(r for r in payload["comics"] if r["gixen_item_id"] == "99")
    assert result["manual_series_flags_cleared"] == 1
    assert row["needs_manual_series_canonical"] is False
    # Narrowest range containing 2015 wins, exactly as the plain (unscoped)
    # lookup would pick — no publisher redirect needed or applied.
    assert row["series_name"] == "Spawn (2012 - Present)"


def test_reresolve_manual_series_publisher_conflict_rescopes_to_correct_volume_bui586(tmp_path):
    """BUI-564's own real-world example: `spawn` carries an Image volume AND a
    foreign (Kamite) licensed edition our own record-win push created. The
    plain era lookup prefers the narrower Kamite range and would misfile the
    win there; the win's OWN publisher (Image Comics) demonstrably conflicts
    with that answer, so the rescope must redirect it to the Image volume."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    image_volume = make_agent_win_row(
        publisher="Image Comics",
        series="Spawn (1992 - Present)",
        full_title="Spawn #1",
        release_date="1992-05-01",
        gixen_item_id=None,
        pushed="2024-01-01T00:00:00.000000Z",
    )
    image_volume["source"] = "locg_export"
    kamite_volume = make_agent_win_row(
        publisher="Kamite",
        series="Spawn (2012 - Present)",
        full_title="Spawn #224",
        release_date="2012-10-01",
        gixen_item_id=None,
        pushed="2024-01-01T00:00:00.000000Z",
    )
    kamite_volume["source"] = "locg_export"
    flagged = make_agent_win_row(
        publisher="Image Comics",
        series="Spawn",  # bare masthead — unresolved
        full_title="Spawn #250",
        release_date="2015-01-01",
        gixen_item_id="99",
        needs_manual_series=True,
        pushed=None,
    )

    def add_rows(payload):
        payload["comics"].extend([image_volume, kamite_volume, flagged])

    cache.apply(add_rows, command="pre-import")

    xlsx = tmp_path / "reexport.xlsx"
    _build_export_xlsx(xlsx, [{
        "publisher": "DC Comics", "series": "Batman (Vol. 3) (2016 - Present)",
        "full_title": "Batman #1", "release_date": "2016-06-15",
    }])

    result = import_xlsx(xlsx, cache)
    payload = cache.load()

    row = next(r for r in payload["comics"] if r["gixen_item_id"] == "99")
    assert result["manual_series_flags_cleared"] == 1
    assert row["needs_manual_series_canonical"] is False
    assert row["series_name"] == "Spawn (1992 - Present)", (
        "must land on the Image volume, not the narrower Kamite one, because "
        "the win's own publisher demonstrably conflicts with Kamite"
    )


def test_reresolve_manual_series_conflict_with_empty_scoped_pool_falls_open_bui586(tmp_path):
    """The ONLY known volume under the key disagrees with the win's publisher
    (no sibling volume exists that agrees), so `publisher_scoped_volume_candidates`
    fails open per-key and hands back the FULL unfiltered list — the rescope is
    attempted, but it can only reproduce the plain lookup's answer. The row must
    still clear its flag and pick up that (unscoped) answer; a conflict being
    detected must never blank or downgrade an existing resolution."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    kamite_volume = make_agent_win_row(
        publisher="Kamite",
        series="Spawn (2012 - Present)",
        full_title="Spawn #224",
        release_date="2012-10-01",
        gixen_item_id=None,
        pushed="2024-01-01T00:00:00.000000Z",
    )
    kamite_volume["source"] = "locg_export"
    flagged = make_agent_win_row(
        publisher="Image Comics",  # conflicts with the only known volume
        series="Spawn",
        full_title="Spawn #250",
        release_date="2015-01-01",
        gixen_item_id="99",
        needs_manual_series=True,
        pushed=None,
    )

    def add_rows(payload):
        payload["comics"].extend([kamite_volume, flagged])

    cache.apply(add_rows, command="pre-import")

    xlsx = tmp_path / "reexport.xlsx"
    _build_export_xlsx(xlsx, [{
        "publisher": "DC Comics", "series": "Batman (Vol. 3) (2016 - Present)",
        "full_title": "Batman #1", "release_date": "2016-06-15",
    }])

    result = import_xlsx(xlsx, cache)
    payload = cache.load()

    row = next(r for r in payload["comics"] if r["gixen_item_id"] == "99")
    assert result["manual_series_flags_cleared"] == 1
    assert row["needs_manual_series_canonical"] is False
    # Same answer the plain (unscoped) lookup would have produced — a
    # detected conflict with nowhere safe to redirect to is a no-op, not a
    # refusal.
    assert row["series_name"] == "Spawn (2012 - Present)"


def test_reresolve_manual_series_null_publisher_degrades_to_todays_answer_bui586(tmp_path):
    """A win row with a null `publisher_name` (the BUI-458 backfill class,
    explicitly out of scope for BUI-586) never reaches the rescope branch at
    all — `series_publisher_conflicts` fails open on a missing publisher, so
    the row degrades to exactly today's (publisher-blind) answer."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    image_volume = make_agent_win_row(
        publisher="Image Comics",
        series="Spawn (1992 - Present)",
        full_title="Spawn #1",
        release_date="1992-05-01",
        gixen_item_id=None,
        pushed="2024-01-01T00:00:00.000000Z",
    )
    image_volume["source"] = "locg_export"
    kamite_volume = make_agent_win_row(
        publisher="Kamite",
        series="Spawn (2012 - Present)",
        full_title="Spawn #224",
        release_date="2012-10-01",
        gixen_item_id=None,
        pushed="2024-01-01T00:00:00.000000Z",
    )
    kamite_volume["source"] = "locg_export"
    flagged = make_agent_win_row(
        series="Spawn",
        full_title="Spawn #250",
        release_date="2015-01-01",
        gixen_item_id="99",
        needs_manual_series=True,
        pushed=None,
    )
    flagged["publisher_name"] = None  # BUI-458 class: null, not missing-key

    def add_rows(payload):
        payload["comics"].extend([image_volume, kamite_volume, flagged])

    cache.apply(add_rows, command="pre-import")

    xlsx = tmp_path / "reexport.xlsx"
    _build_export_xlsx(xlsx, [{
        "publisher": "DC Comics", "series": "Batman (Vol. 3) (2016 - Present)",
        "full_title": "Batman #1", "release_date": "2016-06-15",
    }])

    result = import_xlsx(xlsx, cache)
    payload = cache.load()

    row = next(r for r in payload["comics"] if r["gixen_item_id"] == "99")
    assert result["manual_series_flags_cleared"] == 1
    assert row["needs_manual_series_canonical"] is False
    # No publisher on the row -> no trigger -> the plain era lookup's answer,
    # narrowest range wins (Kamite) — the gap BUI-586 knowingly leaves open.
    assert row["series_name"] == "Spawn (2012 - Present)"


def test_duplicate_check_title_key_strips_the_article_word_safely():
    """`Theatre #1` must not become `atre#1` — the article strip runs on the
    spaced form, before whitespace is collapsed."""
    from locg.collection_io import _duplicate_check_title_key as key

    assert key("Infinity Gauntlet #2") == key("The Infinity Gauntlet #2")
    assert key("Dawn Runner #1") == key("Dawnrunner #1")
    assert key("Theatre #1") == "theatre#1"
    # An edition suffix is part of the key: a base issue and its printing are
    # two books, not a duplicate pair.
    assert key("Batman #2") != key("Batman #2 3rd Printing")


# ---------------------------------------------------------------------------
# BUI-554: the identity key manufactured duplicates, and the counter that was
# supposed to catch them had gone vacuous. Two independent defects, one set of
# regressions — plus the holdings that must survive both fixes.
# ---------------------------------------------------------------------------

# --- Must survive: legitimately distinct rows the fixes must NOT collapse ----

def test_three_xmen_volumes_keep_three_identities():
    """`X-Men #17` is legitimately owned three times. The end-year fold must
    not touch what separates them.

    These are the real store's rows (2026-07-28). Note volumes 1 and 2 differ
    ONLY in the `(Vol. N)` token once the end year folds — which is exactly why
    `identity_series_key` must not reuse `_normalize_series_key`, whose job is
    ownership lookup and which strips `(Vol. N)` outright."""
    from locg.collection_cache import make_identity

    rows = [
        {"publisher_name": "Marvel Comics", "series_name": "The X-Men (Vol. 1) (1963 - 1981)",
         "full_title": "The X-Men #17", "release_date": "1965-12-02"},
        {"publisher_name": "Marvel Comics", "series_name": "X-Men (Vol. 2) (1991 - 2001)",
         "full_title": "X-Men #17", "release_date": "1992-12-15"},
        {"publisher_name": "Marvel Comics", "series_name": "X-Men (Vol. 6) (2019 - 2021)",
         "full_title": "X-Men #17", "release_date": "2021-01-27"},
    ]
    assert len({make_identity(r) for r in rows}) == 3


def test_reused_vol_label_is_separated_by_the_start_year():
    """LOCG reuses a `Vol. N` label across genuinely different volumes, so the
    START year is sometimes the only thing telling two series apart.

    Both pairs are live store rows. Stripping the whole `(YYYY - YYYY)` range —
    the obvious reading of "fold the volume end year" — collapses both, which
    is why `identity_series_key` folds the END year only."""
    from locg.collection_cache import identity_series_key as key

    assert key("X-Men (Vol. 2) (1991 - 2001)") != key("X-Men (Vol. 2) (2001 - 2013)")
    assert key("Spawn (1992 - Present)") != key("Spawn (2012 - Present)")
    # ...while the drift class it exists for does collapse.
    assert (key("Absolute Martian Manhunter (2025 - Present)")
            == key("Absolute Martian Manhunter (2025 - 2026)"))
    assert key("Batman (Vol. 3) (2016 - Present)") == key("Batman (Vol. 3) (2016 - 2026)")
    # A name with no range decoration is returned unchanged.
    assert key("Batman (Vol. 3)") == "Batman (Vol. 3)"
    assert key("") == ""


def test_printings_and_similar_titles_are_not_duplicates():
    """A printing is a distinct collectible, and a title that merely contains
    another is a different book. Both must stay out of the duplicate count."""
    from locg.collection_io import _duplicate_check_title_key as key

    assert (key("Batman: The Dark Knight Returns #2")
            != key("Batman: The Dark Knight Returns #2 3rd Printing"))
    assert key("Death of the Silver Surfer #1") != key("Silver Surfer #1")


def test_duplicate_check_spares_multi_volume_and_printing_holdings(tmp_path):
    """End-to-end: the repaired all-pairs check must report ZERO against a
    store holding only legitimately distinct books — three X-Men volumes, a
    printing, and two same-numbered issues of different Silver Surfer titles.

    The all-pairs widening is only safe because the date predicate still
    rules; this is the test that would fail if a future change let a shared
    title key alone flag a duplicate."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    xlsx = tmp_path / "export.xlsx"
    _build_export_xlsx(xlsx, [
        {"publisher": "Marvel Comics", "series": "The X-Men (Vol. 1) (1963 - 1981)",
         "full_title": "The X-Men #17", "release_date": "1965-12-02"},
        {"publisher": "Marvel Comics", "series": "X-Men (Vol. 2) (1991 - 2001)",
         "full_title": "X-Men #17", "release_date": "1992-12-15"},
        {"publisher": "Marvel Comics", "series": "X-Men (Vol. 6) (2019 - 2021)",
         "full_title": "X-Men #17", "release_date": "2021-01-27"},
        {"publisher": "DC Comics", "series": "Batman: The Dark Knight Returns (1986)",
         "full_title": "Batman: The Dark Knight Returns #2", "release_date": "1986-03-25"},
        {"publisher": "DC Comics", "series": "Batman: The Dark Knight Returns (1986)",
         "full_title": "Batman: The Dark Knight Returns #2 3rd Printing",
         "release_date": "1986-03-31"},
        {"publisher": "Marvel Comics", "series": "Silver Surfer (Vol. 3) (1987 - 1998)",
         "full_title": "Silver Surfer #1", "release_date": "1987-04-07"},
        {"publisher": "Marvel Comics", "series": "Death of the Silver Surfer (1990)",
         "full_title": "Death of the Silver Surfer #1", "release_date": "1990-06-05"},
    ])

    result = import_xlsx(xlsx, cache)
    assert result["added"] == 7
    assert result["owned_duplicate_identities"] == 0
    assert len(cache.load()["comics"]) == 7


# --- Cause A: the series end-year relabel (recurs every January) -------------

def test_series_end_year_relabel_updates_instead_of_inserting(tmp_path):
    """LOCG closes out an ongoing volume's decoration the January after it
    ends. Nothing about the book changed, so the re-import must UPDATE the row
    it already has, not add a twin."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    first = tmp_path / "first.xlsx"
    _build_export_xlsx(first, [
        {"publisher": "DC Comics", "series": "Absolute Martian Manhunter (2025 - Present)",
         "full_title": "Absolute Martian Manhunter #2", "release_date": "2025-04-23"},
    ])
    import_xlsx(first, cache)

    second = tmp_path / "second.xlsx"
    _build_export_xlsx(second, [
        {"publisher": "DC Comics", "series": "Absolute Martian Manhunter (2025 - 2026)",
         "full_title": "Absolute Martian Manhunter #2", "release_date": "2025-04-23"},
    ])
    result = import_xlsx(second, cache)

    payload = cache.load()
    assert len(payload["comics"]) == 1, "the relabel must not manufacture a row"
    assert result["added"] == 0
    assert result["updated"] == 1
    assert result["owned_duplicate_identities"] == 0
    assert payload["comics"][0]["series_name"] == "Absolute Martian Manhunter (2025 - 2026)"


def test_different_start_year_volumes_still_insert_separately(tmp_path):
    """The complement of the test above: two volumes of one masthead that
    differ in START year are two books and must both land."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    first = tmp_path / "first.xlsx"
    _build_export_xlsx(first, [
        {"publisher": "Image Comics", "series": "Spawn (1992 - Present)",
         "full_title": "Spawn #1", "release_date": "1992-05-01"},
    ])
    import_xlsx(first, cache)

    second = tmp_path / "second.xlsx"
    _build_export_xlsx(second, [
        {"publisher": "Image Comics", "series": "Spawn (2012 - Present)",
         "full_title": "Spawn #1", "release_date": "2012-08-01"},
    ])
    result = import_xlsx(second, cache)

    assert result["added"] == 1
    assert len(cache.load()["comics"]) == 2


# --- Cause B: the release-date convention change ----------------------------

def test_release_date_drift_merges_instead_of_inserting(tmp_path):
    """A run re-catalogued from cover date to on-sale date. `release_date` sits
    in BOTH the identity key and the rename detector, so before BUI-554 both
    guards missed together and the row inserted as a twin."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    first = tmp_path / "first.xlsx"
    _build_export_xlsx(first, [
        {"publisher": "DC Comics", "series": "Crisis on Infinite Earths (1985 - 1986)",
         "full_title": "Crisis on Infinite Earths #1", "release_date": "1985-01-03"},
    ])
    import_xlsx(first, cache)

    second = tmp_path / "second.xlsx"
    _build_export_xlsx(second, [
        {"publisher": "DC Comics", "series": "Crisis on Infinite Earths (1985 - 1986)",
         "full_title": "Crisis on Infinite Earths #1", "release_date": "1984-12-11"},
    ])
    result = import_xlsx(second, cache)

    payload = cache.load()
    assert len(payload["comics"]) == 1
    assert result["added"] == 0
    assert result["release_date_drift_merged"] == 1
    assert result["owned_duplicate_identities"] == 0
    assert payload["comics"][0]["release_date"] == "1984-12-11", "re-keyed to LOCG's value"
    # Idempotent: the merged row now matches exactly, so a repeat is a no-op.
    repeat = import_xlsx(second, cache)
    assert repeat["added"] == 0 and repeat["release_date_drift_merged"] == 0
    assert len(cache.load()["comics"]) == 1


def test_release_date_drift_merge_reports_behavioral_drift(tmp_path):
    """A drift merge overwrites LOCG's columns, and those include the
    user-managed ones. Whichever path a row is matched on, a user-edited
    `condition`/`notes`/`grading` that LOCG overwrites has to be reported —
    otherwise the newest match path is the one place hand-entered data can
    vanish unaudited."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    first = tmp_path / "first.xlsx"
    _build_export_xlsx(first, [
        {"publisher": "DC Comics", "series": "Crisis on Infinite Earths (1985 - 1986)",
         "full_title": "Crisis on Infinite Earths #1", "release_date": "1985-01-03"},
    ])
    import_xlsx(first, cache)

    def hand_edit(payload):
        payload["comics"][0]["notes"] = "signed at con"

    cache.apply(hand_edit, command="hand-edit")

    second = tmp_path / "second.xlsx"
    _build_export_xlsx(second, [
        {"publisher": "DC Comics", "series": "Crisis on Infinite Earths (1985 - 1986)",
         "full_title": "Crisis on Infinite Earths #1", "release_date": "1984-12-11"},
    ])
    result = import_xlsx(second, cache)

    assert result["release_date_drift_merged"] == 1
    assert result["behavioral_drift_count"] == 1, (
        "the drift path must audit an overwritten user column like every other "
        "match path"
    )
    records = [
        json.loads(line)
        for line in (tmp_path / "import-history.jsonl").read_text().splitlines()
        if line.strip()
    ]
    drift = [r for r in records if r["type"] == "behavioral_drift"]
    assert len(drift) == 1
    assert drift[0]["details"]["columns_changed"] == ["notes"]


def test_full_title_rename_reports_behavioral_drift(tmp_path):
    """The same invariant on the rename path, which used to compare user
    columns AFTER the overwrite had already made them equal — so its
    `changed` list was always empty and this audit was never written."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    first = tmp_path / "first.xlsx"
    _build_export_xlsx(first, [
        {"publisher": "Marvel Comics", "series": "Nova (Vol. 1) (1976 - 1979)",
         "full_title": "Nova #1", "release_date": "1976-05-01"},
    ])
    import_xlsx(first, cache)

    def hand_edit(payload):
        payload["comics"][0]["notes"] = "reader copy"

    cache.apply(hand_edit, command="hand-edit")

    second = tmp_path / "second.xlsx"
    _build_export_xlsx(second, [
        {"publisher": "Marvel Comics", "series": "Nova (Vol. 1) (1976 - 1979)",
         "full_title": "Nova #1 Facsimile Edition", "release_date": "1976-05-01"},
    ])
    result = import_xlsx(second, cache)

    payload = cache.load()
    assert payload["comics"][0]["previous_full_title"] == "Nova #1", "rename path taken"
    assert result["behavioral_drift_count"] == 1


def test_release_date_drift_beyond_tolerance_stays_a_separate_book(tmp_path):
    """The tolerance is the reconciler's own. Two eras of one title are two
    books, and no amount of title agreement may merge them."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    first = tmp_path / "first.xlsx"
    _build_export_xlsx(first, [
        {"publisher": "Marvel Comics", "series": "Fantastic Four (Vol. 1) (1961 - 2012)",
         "full_title": "Fantastic Four #1", "release_date": "1961-08-08"},
    ])
    import_xlsx(first, cache)

    second = tmp_path / "second.xlsx"
    _build_export_xlsx(second, [
        {"publisher": "Marvel Comics", "series": "Fantastic Four (Vol. 1) (1961 - 2012)",
         "full_title": "Fantastic Four #1", "release_date": "1996-11-13"},
    ])
    result = import_xlsx(second, cache)

    assert result["added"] == 1
    assert result["release_date_drift_merged"] == 0
    assert len(cache.load()["comics"]) == 2


# --- Not cause C: the foreign licensed edition (BUI-559) --------------------

def test_licensed_edition_twin_stays_two_books(tmp_path):
    """A DC row and its Panini twin are two books, not one row that drifted.

    BUI-559 was filed as a third drift class — "same series, same issue, same
    release date, publisher relabelled" — and the store disagrees. All five
    live pairs put Panini 147-211 days AFTER DC, always in that direction:
    Panini DC Italia's Italian edition, its own LOCG catalog entry, its own
    on-sale date. Nothing here may merge them, and the shape below is the one
    that actually occurs, so this is the case to hold.
    """
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    first = tmp_path / "first.xlsx"
    _build_export_xlsx(first, [
        {"publisher": "DC Comics", "series": "Absolute Flash (2025 - Present)",
         "full_title": "Absolute Flash #10", "release_date": "2025-12-17",
         "price_paid": "1.25", "date_purchased": "2026-06-06"},
    ])
    import_xlsx(first, cache)

    # The next export carries BOTH: the round-trip pushed our win onto the
    # Italian entry, so LOCG now reports the book owned twice. Our import must
    # mirror that faithfully — the duplicate is LOCG's, and folding it here
    # would neither be true nor survive the next export.
    second = tmp_path / "second.xlsx"
    _build_export_xlsx(second, [
        {"publisher": "DC Comics", "series": "Absolute Flash (2025 - Present)",
         "full_title": "Absolute Flash #10", "release_date": "2025-12-17",
         "price_paid": "1.25", "date_purchased": "2026-06-06"},
        {"publisher": "Panini Comics", "series": "Absolute Flash (2025 - Present)",
         "full_title": "Absolute Flash #10", "release_date": "2026-06-04",
         "price_paid": "1.25", "date_purchased": "2026-06-06"},
    ])
    result = import_xlsx(second, cache)

    payload = cache.load()
    assert len(payload["comics"]) == 2, "the licensed edition is its own book"
    assert result["added"] == 1
    assert result["release_date_drift_merged"] == 0
    assert sorted(r["publisher_name"] for r in payload["comics"]) == [
        "DC Comics", "Panini Comics",
    ]
    # An identical price_paid + date_purchased is NOT license to merge: it is
    # the round-trip fingerprint of one push, which is exactly how the wrong
    # entry got owned in the first place.
    assert result["owned_duplicate_identities"] == 0, (
        "known blind spot, asserted so it is not mistaken for a clean store: "
        "the counter needs `_release_dates_compatible_either_way`, and a "
        "year-crossing 169-day offset fails it. Three of the five live pairs "
        "are invisible to the sync's duplicate hard-stop for this reason; the "
        "two BUI-556 cleaned were only visible because they happened to land "
        "in one calendar year."
    )


def test_same_year_licensed_twin_is_the_only_kind_the_counter_sees(tmp_path):
    """Why BUI-559 counted two pairs when the store held five.

    `owned_duplicate_identities` groups by title and then needs
    `_release_dates_compatible_either_way` on the pair. A licensed edition
    trails by ~5-7 months, so whether the operator ever hears about it turns on
    the accident of which calendar year that lands in: the two `#3` pairs stayed
    inside 2025 and were reported (and hand-cleaned by BUI-556), while
    `Absolute Flash #10`, `Absolute Green Lantern #8` and `#9` cross into the
    next year, fail the predicate, and are still owned twice with the sync
    reporting clean.

    Also the non-vacuity guard for the `== 0` above: the group demonstrably
    forms, so that zero is the date predicate's ruling and not an empty bucket.
    """
    from locg.collection_io import _duplicate_check_title_key, import_xlsx

    assert _duplicate_check_title_key("Absolute Flash #3") == "absoluteflash#3"

    cache = make_cache(tmp_path)
    first = tmp_path / "first.xlsx"
    _build_export_xlsx(first, [
        {"publisher": "DC Comics", "series": "Absolute Flash (2025 - Present)",
         "full_title": "Absolute Flash #3", "release_date": "2025-05-21"},
    ])
    import_xlsx(first, cache)

    second = tmp_path / "second.xlsx"
    _build_export_xlsx(second, [
        {"publisher": "DC Comics", "series": "Absolute Flash (2025 - Present)",
         "full_title": "Absolute Flash #3", "release_date": "2025-05-21"},
        {"publisher": "Panini Comics", "series": "Absolute Flash (2025 - Present)",
         "full_title": "Absolute Flash #3", "release_date": "2025-12-18"},
    ])
    result = import_xlsx(second, cache)

    assert len(cache.load()["comics"]) == 2, "still two books; only the report differs"
    assert result["owned_duplicate_identities"] == 1


def test_publisher_only_relabel_is_still_uncovered(tmp_path):
    """The drift class BUI-559 *described* — publisher moves, everything else
    held — is genuinely uncovered, and stays that way because it has never
    happened: zero rows in the live store or the pre-BUI-556 backup have this
    shape. Pinned so the gap is executable rather than re-derived, not because
    inserting is the desired end state. A future pass that merges this should
    change this test, having first shown the shape occurs.
    """
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    first = tmp_path / "first.xlsx"
    _build_export_xlsx(first, [
        {"publisher": "DC Comics", "series": "Absolute Flash (2025 - Present)",
         "full_title": "Absolute Flash #3", "release_date": "2025-05-21"},
    ])
    import_xlsx(first, cache)

    second = tmp_path / "second.xlsx"
    _build_export_xlsx(second, [
        {"publisher": "Panini Comics", "series": "Absolute Flash (2025 - Present)",
         "full_title": "Absolute Flash #3", "release_date": "2025-05-21"},
    ])
    result = import_xlsx(second, cache)

    assert result["added"] == 1
    assert result["release_date_drift_merged"] == 0, (
        "the date detector must not reach across publishers either"
    )
    assert len(cache.load()["comics"]) == 2


def test_date_drift_never_folds_two_export_rows_into_one(tmp_path):
    """An export row that matches EXACTLY owns its row. A second export row
    that only drifts onto it must insert instead of overwriting — folding it in
    would drop a book LOCG says exists (the R11 direction)."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    first = tmp_path / "first.xlsx"
    _build_export_xlsx(first, [
        {"publisher": "DC Comics", "series": "Crisis on Infinite Earths (1985 - 1986)",
         "full_title": "Crisis on Infinite Earths #1", "release_date": "1985-01-03"},
    ])
    import_xlsx(first, cache)

    # Same export now carries BOTH dates. The exact match must win, and the
    # drifted row must land as its own row rather than clobbering it.
    second = tmp_path / "second.xlsx"
    _build_export_xlsx(second, [
        {"publisher": "DC Comics", "series": "Crisis on Infinite Earths (1985 - 1986)",
         "full_title": "Crisis on Infinite Earths #1", "release_date": "1984-12-11"},
        {"publisher": "DC Comics", "series": "Crisis on Infinite Earths (1985 - 1986)",
         "full_title": "Crisis on Infinite Earths #1", "release_date": "1985-01-03"},
    ])
    result = import_xlsx(second, cache)

    payload = cache.load()
    dates = sorted(r["release_date"] for r in payload["comics"])
    assert dates == ["1984-12-11", "1985-01-03"], "both export rows must survive"
    assert result["release_date_drift_merged"] == 0


def test_date_drift_pass_never_claims_a_pending_win(tmp_path):
    """A pending win is Phase 1's business. Phase 1 judges it with era and
    print/variant evidence the drift pass does not carry, so the drift pass
    must leave every unpushed win alone."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    win = make_agent_win_row(
        publisher="Marvel Comics",
        series="Uncanny X-Men (Vol. 1) (1963 - 2011)",
        full_title="Uncanny X-Men #179",
        release_date="1983-12-06",
        needs_manual_variant=True,
    )

    def add_win(payload):
        payload["comics"].append(win)

    cache.apply(add_win, command="pre-import")

    xlsx = tmp_path / "export.xlsx"
    _build_export_xlsx(xlsx, [
        {"publisher": "Marvel Comics", "series": "Uncanny X-Men (Vol. 1) (1963 - 2011)",
         "full_title": "Uncanny X-Men #179", "release_date": "1983-11-15"},
    ])
    result = import_xlsx(xlsx, cache)

    assert result["release_date_drift_merged"] == 0, (
        "the drift pass must not adjudicate a win"
    )


# --- The counter: it must be able to fail, whatever the rows are labelled ----

@pytest.mark.parametrize("sources", [
    ("agent_win", "locg_export"),
    ("locg_export", "locg_export"),
    ("agent_win", "agent_win"),
])
def test_duplicate_check_is_blind_to_the_source_label(tmp_path, sources):
    """The guard test for a guard that went vacuous.

    The shipped check partitioned owned rows on `source` and reported only a
    win-vs-export collision. A collection sync round-tripped every win back as
    an export row, draining that partition to zero, and the cross-product
    silently began iterating nothing — 0 reported against 60 real collisions,
    with the healthy reading and the blind reading the same number.

    Parametrizing over the labels is what makes this test unable to go vacuous
    the same way: whichever partition empties next, one of these cases still
    forces the check to fire."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)

    def seed(payload):
        for i, src in enumerate(sources):
            payload["comics"].append({
                "publisher_name": "DC Comics",
                "series_name": "Crisis on Infinite Earths (1985 - 1986)",
                "full_title": "Crisis on Infinite Earths #1",
                # Distinct dates keep the two rows distinct identities; they are
                # 23 days apart, so the reconciler's own predicate calls them
                # the same book.
                "release_date": ["1985-01-03", "1984-12-11"][i],
                "in_collection": 1,
                "in_wish_list": 0,
                "source": src,
                "pushed_to_locg_at": "2026-01-01T00:00:00Z",
                "local_added_at": "2026-01-01T00:00:00Z",
                "local_added_seq": i,
            })

    cache.apply(seed, command="pre-import")

    # An unrelated export, so the merge phases leave the seeded pair intact.
    xlsx = tmp_path / "export.xlsx"
    _build_export_xlsx(xlsx, [
        {"publisher": "Marvel Comics", "series": "Daredevil (Vol. 1) (1964 - 2011)",
         "full_title": "Daredevil #181", "release_date": "1982-01-05"},
    ])
    result = import_xlsx(xlsx, cache)

    assert result["owned_duplicate_identities"] == 1, (
        f"the duplicate must be reported whatever the rows are labelled {sources}"
    )
    assert any("owned TWICE" in w for w in result["warnings"])


def test_duplicate_check_announces_when_it_cannot_check(tmp_path):
    """A check that has lost the ability to fail is itself news. Against a
    store with no owned rows the counter reads 0 for the wrong reason, and
    that 0 must not silently satisfy the sync's hard-stop."""
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    xlsx = tmp_path / "export.xlsx"
    _build_export_xlsx(xlsx, [
        {"publisher": "Marvel Comics", "series": "Daredevil (Vol. 1) (1964 - 2011)",
         "full_title": "Daredevil #181", "release_date": "1982-01-05",
         "in_collection": 0, "in_wish_list": 1},
    ])
    result = import_xlsx(xlsx, cache)

    assert result["owned_duplicate_identities"] == 0
    assert any("VACUOUS" in w for w in result["warnings"]), (
        "a 0 that means 'unable to check' must say so"
    )


# ---------------------------------------------------------------------------
# BUI-563: cross-edition owned twins — ADVISORY, never a sync hard stop.
#
# A foreign licensed edition trails its US original by 147-211 days, an order of
# magnitude past _COVER_TO_ONSALE_MAX_DAYS, so the owned_duplicate_identities
# date predicate structurally cannot see these pairs. They get their own counter
# rather than being folded into the hard stop: the operator has no local remedy
# (LOCG re-emits a deleted row, and clearing the ownership runs the BUI-122
# In Collection=0 data-loss path), so blocking would stop every sync forever.
# ---------------------------------------------------------------------------

def _import_two_owned(tmp_path, left: dict, right: dict):
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    xlsx = tmp_path / "reexport.xlsx"
    _build_export_xlsx(xlsx, [left, right])
    return import_xlsx(xlsx, cache)


def test_cross_edition_twin_is_advisory_not_a_hard_stop(tmp_path):
    """The measured Panini shape: same issue, publishers differ, release dates
    169 days apart, and the SAME price_paid + date_purchased — the round-trip
    fingerprint proving our own push created the foreign row."""
    result = _import_two_owned(
        tmp_path,
        {
            "publisher": "DC Comics",
            "series": "Absolute Flash (2025 - Present)",
            "full_title": "Absolute Flash #10",
            "release_date": "2025-12-17",
            "price_paid": "1.25",
            "date_purchased": "2026-06-06",
        },
        {
            "publisher": "Panini Comics",
            "series": "Absolute Flash (2025 - Present)",
            "full_title": "Absolute Flash #10",
            "release_date": "2026-06-04",
            "price_paid": "1.25",
            "date_purchased": "2026-06-06",
        },
    )

    assert result["owned_duplicate_identities_cross_edition"] == 1
    # The whole point of a separate counter: the sync's hard stop is untouched,
    # so the sync still runs.
    assert result["owned_duplicate_identities"] == 0
    warning = next(w for w in result["warnings"] if "across editions" in w)
    assert "ADVISORY (not a sync blocker)" in warning
    assert "absoluteflash#10" in warning
    assert "In Collection=0" in warning, "must name the BUI-122 trap it forbids"


def test_cross_edition_check_ignores_a_masthead_reused_by_another_publisher(tmp_path):
    """Marvel's `The Transformers` (1984) #13 and Image's `Transformers` (2023)
    #13 are two genuinely different books, legitimately owned side by side. A
    bare publisher-differs test flags them — measured over the live store, this
    exact pair is 2 of its 8 hits — so the round-trip fingerprint is required,
    and it drops them."""
    result = _import_two_owned(
        tmp_path,
        {
            "publisher": "Marvel Comics",
            "series": "The Transformers (1984 - 1991)",
            "full_title": "Transformers #13",
            "release_date": "1985-10-22",
        },
        {
            "publisher": "Image Comics",
            "series": "Transformers (2023 - Present)",
            "full_title": "Transformers #13",
            "release_date": "2024-10-09",
        },
    )

    assert result["owned_duplicate_identities_cross_edition"] == 0
    assert result["owned_duplicate_identities"] == 0


def test_cross_edition_check_requires_both_price_and_purchase_date(tmp_path):
    """A shared price with a DIFFERENT date_purchased is two separate buys that
    happened to cost the same, not one purchase filed twice."""
    result = _import_two_owned(
        tmp_path,
        {
            "publisher": "Marvel Comics",
            "series": "X-Men (Vol. 2) (1991 - 2001)",
            "full_title": "X-Men #59",
            "release_date": "1996-12-01",
            "price_paid": "56.00",
            "date_purchased": "2026-06-12",
        },
        {
            "publisher": "Panini Comics",
            "series": "X-Men (Vol. 2) (2001 - 2013)",
            "full_title": "X-Men #59",
            "release_date": "2005-11-24",
            "price_paid": "56.00",
            "date_purchased": "2026-02-01",
        },
    )

    assert result["owned_duplicate_identities_cross_edition"] == 0


def test_cross_edition_check_ignores_a_same_publisher_pair(tmp_path):
    """Two eras of one masthead under ONE publisher are the ordinary
    two-volumes case, not a licensed edition — whatever else is true of them."""
    result = _import_two_owned(
        tmp_path,
        {
            "publisher": "Marvel Comics",
            "series": "The X-Men (Vol. 1) (1963 - 1981)",
            "full_title": "X-Men #128",
            "release_date": "1979-09-18",
            "price_paid": "10.00",
            "date_purchased": "2026-06-06",
        },
        {
            "publisher": "Marvel Comics",
            "series": "X-Men (Vol. 2) (2001 - 2013)",
            "full_title": "X-Men #128",
            "release_date": "2002-01-15",
            "price_paid": "10.00",
            "date_purchased": "2026-06-06",
        },
    )

    assert result["owned_duplicate_identities_cross_edition"] == 0


def test_cross_edition_counter_is_disjoint_from_the_hard_stop(tmp_path):
    """A title the hard stop already reports must not be counted twice — the two
    numbers are read independently in the sync report."""
    result = _import_two_owned(
        tmp_path,
        {
            "publisher": "DC Comics",
            "series": "Absolute Flash (2025 - Present)",
            "full_title": "Absolute Flash #10",
            "release_date": "2025-12-17",
            "price_paid": "1.25",
            "date_purchased": "2026-06-06",
        },
        {
            # Same YEAR, so _release_dates_compatible_either_way accepts it and
            # the existing hard stop owns this pair.
            "publisher": "Panini Comics",
            "series": "Absolute Flash (2025 - Present)",
            "full_title": "Absolute Flash #10",
            "release_date": "2025-12-30",
            "price_paid": "1.25",
            "date_purchased": "2026-06-06",
        },
    )

    assert result["owned_duplicate_identities"] == 1
    assert result["owned_duplicate_identities_cross_edition"] == 0


# ---------------------------------------------------------------------------
# BUI-564: publisher-scoped volume candidates, so a foreign licensed edition of
# the same masthead cannot capture a win at resolution time.
# ---------------------------------------------------------------------------

_SPAWN_CANDIDATES = {"spawn": ["Spawn (1992 - Present)", "Spawn (2012 - Present)"]}
_SPAWN_PUBLISHERS = {
    "Spawn (1992 - Present)": {"image"},
    "Spawn (2012 - Present)": {"kamite"},
}


def test_unscoped_resolution_picks_the_foreign_volume():
    """Documents the defect this fix exists for. `_best_volume_by_year` prefers
    the NARROWEST range containing the year, so the Kamite volume beats Image's
    for every Spawn win from 2012 on — measured live on the 2026-07-28 store."""
    from locg.collection_cache import resolve_series_for_win

    assert resolve_series_for_win(
        "spawn", "224", 2012, {}, _SPAWN_CANDIDATES
    ) == "Spawn (2012 - Present)"


def test_publisher_scoping_redirects_a_win_to_its_own_publishers_volume():
    from locg.collection_cache import resolve_series_for_win
    from locg.collection_io import publisher_scoped_volume_candidates

    scoped = publisher_scoped_volume_candidates(
        _SPAWN_CANDIDATES, _SPAWN_PUBLISHERS, "Image Comics"
    )
    assert scoped["spawn"] == ["Spawn (1992 - Present)"]
    assert resolve_series_for_win(
        "spawn", "224", 2012, {}, scoped
    ) == "Spawn (1992 - Present)"


def test_publisher_scoping_folds_provider_naming_drift():
    """Metron says "Image", LOCG says "Image Comics" — BUI-548's exact trap.
    _normalize_publisher must absorb it, or the scoping would drop the RIGHT
    volume and make things worse."""
    from locg.collection_io import publisher_scoped_volume_candidates

    scoped = publisher_scoped_volume_candidates(
        _SPAWN_CANDIDATES, _SPAWN_PUBLISHERS, "Image"
    )
    assert scoped["spawn"] == ["Spawn (1992 - Present)"]


def test_publisher_scoping_fails_open_when_nothing_matches():
    """A publisher no local volume names (an imprint, a rebrand) must NOT empty
    the pool — that would stop the win resolving. The key keeps its full list
    and behavior is exactly today's."""
    from locg.collection_io import publisher_scoped_volume_candidates

    scoped = publisher_scoped_volume_candidates(
        _SPAWN_CANDIDATES, _SPAWN_PUBLISHERS, "Dark Horse Comics"
    )
    assert scoped["spawn"] == _SPAWN_CANDIDATES["spawn"]


def test_publisher_scoping_keeps_a_volume_with_no_known_publisher():
    from locg.collection_io import publisher_scoped_volume_candidates

    scoped = publisher_scoped_volume_candidates(
        _SPAWN_CANDIDATES, {"Spawn (2012 - Present)": {"kamite"}}, "Image Comics"
    )
    assert scoped["spawn"] == ["Spawn (1992 - Present)"]


def test_series_publisher_conflicts_needs_positive_disagreement():
    """The trigger fires only on evidence. An unknown publisher on either side
    is not a conflict — treating it as one would rescope wins whose volume was
    never in question."""
    from locg.collection_io import series_publisher_conflicts

    assert series_publisher_conflicts(
        "Spawn (2012 - Present)", "Image Comics", _SPAWN_PUBLISHERS
    )
    assert not series_publisher_conflicts(
        "Spawn (1992 - Present)", "Image Comics", _SPAWN_PUBLISHERS
    )
    # Volume LOCG left publisher-less.
    assert not series_publisher_conflicts("Spawn (1999)", "Image Comics", _SPAWN_PUBLISHERS)
    # Metron had no publisher for the issue.
    assert not series_publisher_conflicts(
        "Spawn (2012 - Present)", "", _SPAWN_PUBLISHERS
    )


def test_build_series_publishers_reads_only_locg_export_rows():
    """R61: an agent_win row's publisher is our own guess and must never define
    what a LOCG volume's publisher is."""
    from locg.collection_io import build_series_publishers

    payload = {"comics": [
        {"source": "locg_export", "series_name": "Spawn (1992 - Present)",
         "publisher_name": "Image Comics"},
        {"source": "agent_win", "series_name": "Spawn (2012 - Present)",
         "publisher_name": "Image Comics"},
        {"source": "locg_export", "series_name": "Spawn (2012 - Present)",
         "publisher_name": "Kamite"},
        {"source": "locg_export", "series_name": "Spawn (1992 - Present)",
         "publisher_name": None},
    ]}

    assert build_series_publishers(payload) == {
        "Spawn (1992 - Present)": {"image"},
        "Spawn (2012 - Present)": {"kamite"},
    }
